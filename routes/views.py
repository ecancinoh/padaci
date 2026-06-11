import json
import math
import uuid
import traceback
import ast
import re
import hashlib
from io import BytesIO
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import URLError
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse
from django.db.models import Q, Max
from django.db import transaction
from django.db.utils import OperationalError, ProgrammingError
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from .models import (
    RutaDia,
    ParadaRuta,
    Entrega,
    EntregaPago,
    FacturaOCRLinea,
    ClienteOCRAlias,
    ParadaFalabellaMeta,
    ParadaUbicacionCandidata,
)
from .forms import (
    RutaDiaForm,
    EntregaForm,
    EntregaEstadoForm,
    RutaFalabellaExcelForm,
    build_entrega_pago_formset,
)
from clients.models import Cliente
from companies.models import Empresa


def _write_routes_log(filename, payload):
    base_dir = Path(__file__).resolve().parents[1]
    candidates = [
        base_dir / 'tmp' / filename,
        base_dir / filename,
        Path('/tmp') / filename,
    ]
    for path in candidates:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open('a', encoding='utf-8') as f:
                f.write(payload + '\n')
            return
        except Exception:
            continue


FALABELLA_SESSION_PREFIX = 'falabella_upload_'
VI_REGION_ALIASES = [
    "Region de O'Higgins",
    'Region del Libertador General Bernardo OHiggins',
    'Sexta Region',
]
COORD_RANGO_RENGO = (-34.4068, -70.8583)
COORD_RANGO_RANCAGUA = (-34.1708, -70.7444)


def _normalize_header(value):
    text = _normalizar(str(value or ''))
    text = text.replace('.', '').replace(':', '').replace('-', ' ').replace('_', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _normalize_patente(value):
    return re.sub(r'[^A-Z0-9]', '', str(value or '').upper()).strip()


def _read_excel_rows(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Falta dependencia openpyxl para leer .xlsx.') from exc

    raw = file_obj.read()
    if not raw:
        raise ValueError('El archivo Excel está vacío.')

    book = load_workbook(filename=BytesIO(raw), data_only=True)

    aliases = {
        'empresa': ['empresa'],
        'patente': ['patente'],
        'direccion': ['direccion', 'direccion cliente', 'direccion de entrega'],
        'localidad': ['localidad', 'comuna', 'ciudad'],
    }

    required = ['empresa', 'patente', 'direccion', 'localidad']

    def _get_col_index(sheet):
        rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not rows:
            return None
        headers_raw = [str(h or '').strip() for h in rows[0]]
        headers_norm = [_normalize_header(h) for h in headers_raw]
        col_index = {}
        for key, names in aliases.items():
            for idx, col_name in enumerate(headers_norm):
                if col_name in names:
                    col_index[key] = idx
                    break
        missing = [name for name in required if name not in col_index]
        if missing:
            return None
        return col_index

    ws = None
    col_index = None

    # 1) Priorizar hojas que explícitamente sean de planificación
    for sheet in book.worksheets:
        title_norm = _normalize_header(sheet.title)
        if 'planificacion' not in title_norm:
            continue
        candidate_cols = _get_col_index(sheet)
        if candidate_cols is not None:
            ws = sheet
            col_index = candidate_cols
            break

    # 2) Fallback: primera hoja que cumpla columnas requeridas
    if ws is None:
        for sheet in book.worksheets:
            candidate_cols = _get_col_index(sheet)
            if candidate_cols is not None:
                ws = sheet
                col_index = candidate_cols
                break

    if ws is None or col_index is None:
        raise ValueError(
            'No se encontró hoja de planificación con columnas requeridas: empresa, patente, direccion y localidad.'
        )

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('La hoja de planificación no contiene datos.')

    parsed_rows = []
    for idx, row in enumerate(rows[1:], start=2):
        empresa = str(row[col_index['empresa']] or '').strip()
        patente = _normalize_patente(row[col_index['patente']])
        direccion = str(row[col_index['direccion']] or '').strip()
        localidad = str(row[col_index['localidad']] or '').strip()

        if not any([empresa, patente, direccion, localidad]):
            continue
        if not empresa or not patente or not direccion:
            continue

        parsed_rows.append({
            'row_number': idx,
            'empresa': empresa,
            'patente': patente,
            'direccion': direccion,
            'localidad': localidad,
        })

    if not parsed_rows:
        raise ValueError('No se encontraron filas válidas en el archivo.')

    empresas = sorted({r['empresa'] for r in parsed_rows if r['empresa']})
    patentes = sorted({r['patente'] for r in parsed_rows if r['patente']})
    return parsed_rows, empresas, patentes


def _falabella_payload_from_session(request, token):
    if not token:
        return None
    return request.session.get(f'{FALABELLA_SESSION_PREFIX}{token}')


def _geocode_nominatim(query=None, limit=4, *, street=None, city=None, state=None):
    params = {
        'format': 'jsonv2',
        'addressdetails': 1,
        'limit': limit,
        'countrycodes': 'cl',
    }
    if street or city or state:
        if street:
            params['street'] = street
        if city:
            params['city'] = city
        if state:
            params['state'] = state
        params['country'] = 'Chile'
    else:
        params['q'] = query or ''

    url = 'https://nominatim.openstreetmap.org/search?' + urlencode(params)
    req = Request(url, headers={'User-Agent': 'PADACI/1.0 (Rutas Falabella)'})
    try:
        with urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except (URLError, TimeoutError, json.JSONDecodeError):
        return []

    candidatos = []
    for idx, item in enumerate(data, start=1):
        try:
            lat = Decimal(str(item.get('lat'))).quantize(Decimal('0.000001'))
            lon = Decimal(str(item.get('lon'))).quantize(Decimal('0.000001'))
        except (InvalidOperation, TypeError):
            continue
        candidatos.append({
            'latitud': lat,
            'longitud': lon,
            'etiqueta': str(item.get('name', '') or '').strip(),
            'direccion_formateada': str(item.get('display_name', '') or '').strip(),
            'score': Decimal(str(item.get('importance', 0) or 0)).quantize(Decimal('0.0001')),
            'orden': idx,
        })
    return candidatos


def _extract_street_number(text):
    match = re.search(r'\b(\d{1,6}[A-Za-z]?)\b', str(text or ''))
    return match.group(1).upper() if match else ''


def _contains_street_number(text, number):
    if not number:
        return True
    haystack = str(text or '').upper()
    return re.search(rf'\b{re.escape(number)}\b', haystack) is not None


def _street_name_without_number(text):
    return re.sub(r'\b\d{1,6}[A-Za-z]?\b', '', str(text or '')).strip()


def _geocode_free_address(direccion, localidad):
    candidatos = _geocode_nominatim(
        limit=10,
        street=direccion or None,
        city=localidad or None,
        state="Region de O'Higgins",
    )
    if not candidatos:
        query = ', '.join(part for part in [direccion, localidad, "Region de O'Higgins", 'Chile'] if part)
        candidatos = _geocode_nominatim(query, limit=10)
    if not candidatos:
        return None

    numero = _extract_street_number(direccion)
    calle_norm = _normalizar(_street_name_without_number(direccion))
    localidad_norm = _normalizar(localidad)
    calle_tokens = [t for t in calle_norm.split() if len(t) > 2]

    def _passes_filters(c, require_number=True, require_locality=True):
        texto = _normalizar(f"{c.get('etiqueta', '')} {c.get('direccion_formateada', '')}")

        if require_number and numero and not _contains_street_number(texto, numero):
            return False

        if calle_tokens:
            # Requiere al menos un token relevante de la calle (ej: lourdes, almagro).
            if not any(token in texto for token in calle_tokens):
                return False

        if require_locality and localidad_norm and localidad_norm not in texto:
            return False

        return True

    candidatos_filtrados = [c for c in candidatos if _passes_filters(c, require_number=True, require_locality=True)]

    # Fallback: si no hay coincidencia exacta, intentar calle/localidad sin exigir número.
    if not candidatos_filtrados:
        candidatos_filtrados = [c for c in candidatos if _passes_filters(c, require_number=False, require_locality=True)]

    # Último fallback: al menos la calle debe coincidir.
    if not candidatos_filtrados:
        candidatos_filtrados = [c for c in candidatos if _passes_filters(c, require_number=False, require_locality=False)]

    if not candidatos_filtrados:
        return None

    def _rank(c):
        texto = _normalizar(f"{c.get('etiqueta', '')} {c.get('direccion_formateada', '')}")
        score = float(c.get('score') or 0)
        if calle_norm and calle_norm in texto:
            score += 1.5
        if numero and _contains_street_number(texto, numero):
            score += 1.0
        if localidad_norm and localidad_norm in texto:
            score += 1.0
        return score

    best = max(candidatos_filtrados, key=_rank)
    return best.get('latitud'), best.get('longitud')


def _falabella_regenerar_candidatos(parada, direccion, localidad):
    cliente = parada.entrega.cliente
    meta, _ = ParadaFalabellaMeta.objects.get_or_create(
        parada=parada,
        defaults={
            'direccion_original': direccion or '',
            'localidad_original': localidad or '',
            'contacto_original': '',
            'estado_direccion': 'pendiente_busqueda',
        },
    )

    meta.direccion_original = direccion or ''
    meta.localidad_original = localidad or ''
    meta.save(update_fields=['direccion_original', 'localidad_original', 'fecha_actualizacion'])

    ParadaUbicacionCandidata.objects.filter(parada=parada).delete()

    candidatos = _geocode_nominatim(
        limit=4,
        street=direccion or None,
        city=localidad or None,
        state="Region de O'Higgins",
    )
    if not candidatos:
        query = ', '.join(part for part in [direccion or '', localidad or '', "Region de O'Higgins", 'Chile'] if part)
        candidatos = _geocode_nominatim(query, limit=4)

    if not candidatos:
        meta.estado_direccion = 'requiere_llamada_cliente'
        meta.save(update_fields=['estado_direccion', 'fecha_actualizacion'])
        return 0, False

    first_id = None
    for cand in candidatos:
        candidate = ParadaUbicacionCandidata.objects.create(
            parada=parada,
            proveedor='nominatim',
            etiqueta=cand['etiqueta'],
            direccion_formateada=cand['direccion_formateada'],
            latitud=cand['latitud'],
            longitud=cand['longitud'],
            score=cand['score'],
            orden=cand['orden'],
        )
        if first_id is None:
            first_id = candidate.pk

    if first_id:
        ParadaUbicacionCandidata.objects.filter(pk=first_id).update(seleccionada=True)
        first = ParadaUbicacionCandidata.objects.get(pk=first_id)
        cliente.latitud = first.latitud
        cliente.longitud = first.longitud
        cliente.save(update_fields=['latitud', 'longitud', 'fecha_actualizacion'])

    meta.estado_direccion = 'candidatos_disponibles'
    meta.save(update_fields=['estado_direccion', 'fecha_actualizacion'])
    return len(candidatos), True


def _falabella_optimizar_con_anclas(ruta):
    paradas = list(ruta.paradas.select_related('entrega__cliente').all())
    con_coords = []
    sin_coords = []
    for p in paradas:
        c = p.entrega.cliente
        if c.latitud is None or c.longitud is None:
            sin_coords.append(p)
            continue
        con_coords.append({
            'parada': p,
            'lat': float(c.latitud),
            'lon': float(c.longitud),
        })

    origen_lat, origen_lon = COORD_RANGO_RENGO
    fin_lat, fin_lon = COORD_RANGO_RANCAGUA
    ordenadas = []
    current_lat, current_lon = origen_lat, origen_lon
    unvisited = list(con_coords)

    while unvisited:
        nearest = min(
            unvisited,
            key=lambda p: (
                _haversine(current_lat, current_lon, p['lat'], p['lon'])
                + (_haversine(p['lat'], p['lon'], fin_lat, fin_lon) * 0.25)
            ),
        )
        unvisited.remove(nearest)
        ordenadas.append(nearest['parada'])
        current_lat, current_lon = nearest['lat'], nearest['lon']

    final = ordenadas + sin_coords
    cambios = [(parada, idx) for idx, parada in enumerate(final, start=1) if parada.orden != idx]
    if cambios:
        with transaction.atomic():
            max_orden_actual = ruta.paradas.aggregate(max_orden=Max('orden')).get('max_orden') or 0
            offset = max_orden_actual + len(final) + 10

            # Fase 1: mover temporalmente para evitar colisiones de unique(ruta, orden)
            for tmp_idx, (parada, _new_idx) in enumerate(cambios, start=1):
                ParadaRuta.objects.filter(pk=parada.pk).update(orden=offset + tmp_idx)

            # Fase 2: aplicar orden final secuencial
            for parada, new_idx in cambios:
                ParadaRuta.objects.filter(pk=parada.pk).update(orden=new_idx)

    return len(ordenadas), len(sin_coords)


class EntregaRutaListView(LoginRequiredMixin, ListView):
    model = Entrega
    template_name = 'deliveries/list.html'
    context_object_name = 'entregas'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('cliente', 'empresa', 'conductor')
        estado = self.request.GET.get('estado', '')
        fecha = self.request.GET.get('fecha', '')
        q = self.request.GET.get('q', '')
        if estado:
            qs = qs.filter(estado=estado)
        if fecha:
            qs = qs.filter(fecha_programada=fecha)
        if q:
            qs = qs.filter(numero_guia__icontains=q) | qs.filter(cliente__nombre__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['estados'] = Entrega.ESTADO_CHOICES
        return ctx


class EntregaRutaCreateView(LoginRequiredMixin, CreateView):
    model = Entrega
    form_class = EntregaForm
    template_name = 'deliveries/form.html'
    success_url = reverse_lazy('routes:entregas_list')

    def form_valid(self, form):
        messages.success(self.request, 'Entrega registrada exitosamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Registrar Entrega'
        return ctx


class EntregaRutaUpdateView(LoginRequiredMixin, UpdateView):
    model = Entrega
    form_class = EntregaForm
    template_name = 'deliveries/form.html'
    success_url = reverse_lazy('routes:entregas_list')

    def form_valid(self, form):
        messages.success(self.request, 'Entrega actualizada correctamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Editar Entrega'
        return ctx


class EntregaRutaDeleteView(LoginRequiredMixin, DeleteView):
    model = Entrega
    template_name = 'deliveries/confirm_delete.html'
    success_url = reverse_lazy('routes:entregas_list')

    def form_valid(self, form):
        messages.success(self.request, 'Entrega eliminada.')
        return super().form_valid(form)


class EntregaRutaDetailView(LoginRequiredMixin, DetailView):
    model = Entrega
    template_name = 'deliveries/detail.html'
    context_object_name = 'entrega'


@login_required
def entregas_eliminar_masiva(request):
    if request.method != 'POST':
        return redirect('routes:entregas_list')
    ids = request.POST.getlist('ids')
    if ids:
        qs = Entrega.objects.filter(pk__in=ids)
        total = qs.count()
        qs.delete()
        messages.success(request, f'{total} entrega(s) eliminada(s) correctamente.')
    else:
        messages.warning(request, 'No se seleccionó ninguna entrega.')
    return redirect('routes:entregas_list')


@login_required
def entrega_actualizar_estado(request, pk):
    entrega = get_object_or_404(Entrega, pk=pk)
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and not url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = None

    if request.method == 'POST':
        form = EntregaEstadoForm(request.POST, request.FILES, instance=entrega)
        pagos_formset = build_entrega_pago_formset(data=request.POST, instance=entrega)
        if form.is_valid() and pagos_formset.is_valid():
            with transaction.atomic():
                obj = form.save(commit=False)
                if obj.estado == 'entregado' and not obj.fecha_entrega:
                    obj.fecha_entrega = timezone.now()

                if form.changed_data or any(f.has_changed() for f in pagos_formset.forms):
                    obj.pago_registrado_por = request.user
                    obj.fecha_registro_pago = timezone.now()

                obj.save()

                pagos_formset.instance = obj
                pagos = pagos_formset.save(commit=False)
                for pago in pagos:
                    pago.registrado_por = request.user
                    pago.save()

                for pago in pagos_formset.deleted_objects:
                    pago.delete()

            messages.success(request, 'Estado y pagos actualizados correctamente.')
            if next_url:
                return redirect(next_url)
            return redirect('routes:entregas_detail', pk=pk)
    else:
        form = EntregaEstadoForm(instance=entrega)
        pagos_formset = build_entrega_pago_formset(instance=entrega)
    return render(
        request,
        'deliveries/actualizar_estado.html',
        {
            'form': form,
            'pagos_formset': pagos_formset,
            'entrega': entrega,
            'next_url': next_url,
        },
    )


@login_required
def eliminar_pago_entrega(request, pk, pago_pk):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Metodo no permitido.'}, status=405)

    entrega = get_object_or_404(Entrega, pk=pk)
    pago = get_object_or_404(EntregaPago, pk=pago_pk, entrega=entrega)
    pago.delete()
    return JsonResponse({'ok': True, 'pago_id': pago_pk})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _haversine(lat1, lon1, lat2, lon2):
    """Distancia en km entre dos coordenadas geográficas."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


def _nearest_neighbor_route(points):
    """
    Algoritmo Nearest Neighbor para TSP (Traveling Salesman Problem).
    Recibe lista de dicts con keys: id, lat, lon.
    Retorna lista ordenada de ids.
    """
    if not points:
        return []
    unvisited = list(points)
    # Punto de inicio: el primero (bodega o inicio del conductor)
    current = unvisited.pop(0)
    route = [current]
    while unvisited:
        nearest = min(unvisited, key=lambda p: _haversine(
            float(current['lat']), float(current['lon']),
            float(p['lat']), float(p['lon']),
        ))
        unvisited.remove(nearest)
        route.append(nearest)
        current = nearest
    return [p['id'] for p in route]


def _resumen_pago_entrega(entrega):
    pagos_qs = getattr(entrega, 'pagos', None)
    if pagos_qs is None:
        total_pagado = 0
    else:
        total_pagado = sum((p.monto or 0) for p in pagos_qs.all())

    return {
        'estado_pago': entrega.estado_pago,
        'estado_pago_display': entrega.get_estado_pago_display(),
        'total_pagado': float(total_pagado or 0),
    }


def _safe_int(value):
    try:
        if value is None or value == '':
            return None
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _safe_total(value):
    """
    Parsea un monto a Decimal. Maneja formatos variados:
    - 50000, 50000.0, 50000.00
    - $50000, $50,000.00
    - 50.000 (formato europeo con . como miles)
    """
    if value is None:
        return Decimal('0')
    
    text = str(value).strip()
    if not text or text == '':
        return Decimal('0')
    
    # Si ya es un número (int/float/Decimal), convertir directamente
    if isinstance(value, (int, float, Decimal)):
        try:
            parsed = Decimal(str(value))
            if parsed < 0:
                return Decimal('0')
            return parsed.quantize(Decimal('1'))
        except (InvalidOperation, ValueError):
            return Decimal('0')
    
    # Remover símbolo de moneda y espacios
    text = text.replace('$', '').replace(' ', '').strip()
    
    if not text:
        return Decimal('0')
    
    # Contar puntos y comas para determinar separadores
    dot_count = text.count('.')
    comma_count = text.count(',')
    
    # Estrategia: el último separador a 2-3 posiciones del final es probablemente decimal
    # De lo contrario, tratar como miles
    if dot_count > 0 and comma_count == 0:
        # Solo hay puntos: si el último punto está a 1-3 posiciones del final, es decimal
        last_dot_pos = text.rfind('.')
        if last_dot_pos > len(text) - 4:
            # Es probable que sea decimal (ej: 50000.00 o 50000.5)
            pass  # Mantener el punto como está
        else:
            # Es probable que sea miles (ej: 50.000.000 o 50.000)
            # Remover todos los puntos
            text = text.replace('.', '')
    elif comma_count > 0 and dot_count == 0:
        # Solo hay comas: si la última coma está a 1-3 posiciones del final, es decimal
        last_comma_pos = text.rfind(',')
        if last_comma_pos > len(text) - 4:
            # Es probable que sea decimal (ej: 50000,00 o 50000,5)
            text = text.replace(',', '.')
        else:
            # Es probable que sea miles (ej: 50,000,000 o 50,000)
            text = text.replace(',', '')
    elif dot_count > 0 and comma_count > 0:
        # Hay ambos: el último separador es probablemente el decimal
        last_dot_pos = text.rfind('.')
        last_comma_pos = text.rfind(',')
        if last_dot_pos > last_comma_pos:
            # Punto es decimal, comas son miles
            text = text.replace(',', '')
        else:
            # Coma es decimal, puntos son miles
            text = text.replace('.', '')
            text = text.replace(',', '.')
    
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal('0')
    
    if parsed < 0:
        return Decimal('0')
    
    return parsed.quantize(Decimal('1'))


def _coerce_facturas_payload(payload):
    if isinstance(payload, dict):
        payload = payload.get('facturas', [])
    if not isinstance(payload, list):
        return []

    filas = []
    for raw in payload:
        if not isinstance(raw, dict):
            continue
        filas.append({
            'documento': str(raw.get('documento', '') or '').strip(),
            'nombre_cliente': str(raw.get('nombre_cliente', '') or '').strip(),
            'direccion_cliente': str(raw.get('direccion_cliente', '') or '').strip(),
            'dia': _safe_int(raw.get('dia')),
            'mes': _safe_int(raw.get('mes')),
            'total': _safe_total(raw.get('total')),
            'transp': str(raw.get('transp', '') or '').strip(),
            'cond_pago': str(raw.get('cond_pago', '') or '').strip(),
            'comuna': str(raw.get('comuna', '') or '').strip(),
        })
    return filas


def _parse_facturas_from_raw_text(raw_text):
    text = (raw_text or '').strip()
    if not text:
        return []

    if text.startswith('```'):
        lines = [line for line in text.splitlines() if not line.strip().startswith('```')]
        text = '\n'.join(lines).strip()

    # Intento 1: JSON directo
    try:
        return _coerce_facturas_payload(json.loads(text))
    except Exception:
        pass

    # Intento 2: encontrar el primer bloque tipo lista JSON en el contenido
    ini = text.find('[')
    fin = text.rfind(']')
    if ini != -1 and fin != -1 and fin > ini:
        block = text[ini:fin + 1]
        try:
            return _coerce_facturas_payload(json.loads(block))
        except Exception:
            try:
                # Gemini a veces usa comillas simples / estilo Python
                return _coerce_facturas_payload(ast.literal_eval(block))
            except Exception:
                pass

    # Intento 3: tabla markdown/pipe
    filas = []
    for line in text.splitlines():
        if '|' not in line:
            continue
        raw_cells = [c.strip() for c in line.split('|') if c.strip()]
        if len(raw_cells) < 5:
            continue
        joined = ' '.join(raw_cells).lower()
        if 'documento' in joined and 'cliente' in joined:
            continue
        if set(joined) <= {'-', ':'}:
            continue

        numero = raw_cells[0] if len(raw_cells) > 0 else ''
        nombre = raw_cells[1] if len(raw_cells) > 1 else ''
        direccion = raw_cells[2] if len(raw_cells) > 2 else ''
        dia = _safe_int(raw_cells[3] if len(raw_cells) > 3 else None)
        mes = _safe_int(raw_cells[4] if len(raw_cells) > 4 else None)
        total = _safe_total(raw_cells[5] if len(raw_cells) > 5 else 0)
        transp = raw_cells[6] if len(raw_cells) > 6 else ''
        cond_pago = raw_cells[7] if len(raw_cells) > 7 else ''
        comuna = raw_cells[8] if len(raw_cells) > 8 else ''
        if nombre or numero:
            filas.append({
                'documento': numero,
                'nombre_cliente': nombre,
                'direccion_cliente': direccion,
                'dia': dia,
                'mes': mes,
                'total': total,
                'transp': transp,
                'cond_pago': cond_pago,
                'comuna': comuna,
            })
    return filas


def _ocr_extract_facturas(image_path):
    """
    Extrae filas de factura desde una imagen usando Gemini en formato JSON.
    Retorna tuple: (filas, texto_raw).
    """
    try:
        from google import genai
        import PIL.Image
        from django.conf import settings

        api_key = settings.GEMINI_API_KEY
        if not api_key or api_key == 'TU_API_KEY_AQUI':
            return ['Error: Configura GEMINI_API_KEY en el archivo .env']

        client = genai.Client(api_key=api_key)
        img = PIL.Image.open(image_path)

        prompt = (
            "Extrae cada fila de la tabla en formato JSON puro (sin markdown). "
            "Responde SOLO un arreglo JSON, donde cada elemento contenga estas claves exactas: "
            "documento, nombre_cliente, direccion_cliente, dia, mes, total, transp, cond_pago, comuna. "
            "Si un dato no aparece claramente, usa cadena vacía. "
            "No inventes datos y conserva el texto original cuando sea posible."
        )

        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[prompt, img],
        )
        text = (response.text or '').strip()
        return _parse_facturas_from_raw_text(text), text

    except Exception as e:
        msg = str(e)
        if '429' in msg or 'RESOURCE_EXHAUSTED' in msg:
            return [], f'Error cuota Gemini: {msg}'
        return [], f'Gemini Error: {msg}'


def _normalizar(texto):
    """Normaliza texto: minúsculas, sin tildes, sin puntuación extra."""
    import unicodedata
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto


def _build_ocr_alias_key(nombre, direccion='', comuna=''):
    raw = '||'.join([
        _normalizar(nombre or ''),
        _normalizar(direccion or ''),
        _normalizar(comuna or ''),
    ])
    if len(raw) <= 240:
        return raw
    digest = hashlib.sha256(raw.encode('utf-8')).hexdigest()
    return f'h:{digest}'


def _cliente_candidatos(nombre, direccion='', comuna='', umbral=0.52, limite=5):
    nombre = (nombre or '').strip()
    if not nombre:
        return []

    nombre_norm = _normalizar(nombre)
    direccion_norm = _normalizar(direccion or '')
    comuna_norm = _normalizar(comuna or '')

    base_qs = Cliente.objects.all().only('id', 'nombre', 'direccion', 'comuna')
    candidatos = []
    for cliente in base_qs:
        nombre_cliente_norm = _normalizar(cliente.nombre or '')
        ratio_nombre = SequenceMatcher(None, nombre_norm, nombre_cliente_norm).ratio()
        ratio_dir = 0.0
        ratio_comuna = 0.0

        if direccion_norm and cliente.direccion:
            ratio_dir = SequenceMatcher(None, direccion_norm, _normalizar(cliente.direccion)).ratio()
        if comuna_norm and cliente.comuna:
            ratio_comuna = SequenceMatcher(None, comuna_norm, _normalizar(cliente.comuna)).ratio()

        score = ratio_nombre
        if direccion_norm:
            score = (score * 0.75) + (ratio_dir * 0.25)
        if comuna_norm:
            score = (score * 0.85) + (ratio_comuna * 0.15)

        if score >= umbral:
            candidatos.append({
                'id': cliente.pk,
                'nombre': cliente.nombre,
                'direccion': cliente.direccion or '',
                'comuna': cliente.comuna or '',
                'score': round(score, 4),
            })

    candidatos.sort(key=lambda c: c['score'], reverse=True)
    return candidatos[:limite]


def _match_clients_from_facturas(facturas, umbral=0.52):
    resultados = []
    no_encontrados = []
    clientes_detectados = {}

    for item in facturas:
        nombre = item.get('nombre_cliente', '').strip()
        direccion = item.get('direccion_cliente', '').strip()
        comuna = item.get('comuna', '').strip()
        alias_key = _build_ocr_alias_key(nombre, direccion, comuna)

        alias = None
        if alias_key and alias_key != '||||':
            alias = (
                ClienteOCRAlias.objects
                .filter(clave_normalizada=alias_key)
                .select_related('cliente')
                .first()
            )

        if alias and alias.bloqueado_por_conflicto:
            candidatos = _cliente_candidatos(nombre, direccion, comuna, umbral=umbral)
            cliente_id = None
            requiere_revision = True
        elif alias and not alias.bloqueado_por_conflicto and alias.cliente_id:
            candidatos = [{
                'id': alias.cliente_id,
                'nombre': alias.cliente.nombre,
                'direccion': alias.cliente.direccion or '',
                'comuna': alias.cliente.comuna or '',
                'score': 1.0,
            }]
            cliente_id = alias.cliente_id
            requiere_revision = False
        else:
            candidatos = _cliente_candidatos(nombre, direccion, comuna, umbral=umbral)
            cliente_id = candidatos[0]['id'] if len(candidatos) == 1 else None
            requiere_revision = not cliente_id

        if cliente_id and candidatos:
            clientes_detectados[cliente_id] = {
                'id': cliente_id,
                'nombre': candidatos[0]['nombre'],
                'comuna': candidatos[0]['comuna'],
            }
        elif nombre:
            no_encontrados.append(nombre)

        resultados.append({
            'documento': item.get('documento', ''),
            'nombre_cliente': nombre,
            'direccion_cliente': direccion,
            'dia': item.get('dia'),
            'mes': item.get('mes'),
            'total': float(item.get('total') or 0),
            'transp': item.get('transp', ''),
            'cond_pago': item.get('cond_pago', ''),
            'comuna': comuna,
            'cliente_id': cliente_id,
            'candidatos': candidatos,
            'requiere_revision': requiere_revision,
        })

    return resultados, list(clientes_detectados.values()), no_encontrados


@login_required
def crear_cliente_rapido_desde_ruta(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    nombre = str(payload.get('nombre', '') or '').strip()
    direccion = str(payload.get('direccion', '') or '').strip()
    comuna = str(payload.get('comuna', '') or '').strip()

    if not nombre:
        return JsonResponse({'error': 'El nombre es obligatorio.'}, status=400)

    cliente = Cliente.objects.create(
        empresa=ruta.empresa,
        nombre=nombre,
        direccion=direccion,
        comuna=comuna,
    )

    return JsonResponse({
        'ok': True,
        'cliente': {
            'id': cliente.pk,
            'nombre': cliente.nombre,
            'direccion': cliente.direccion or '',
            'comuna': cliente.comuna or '',
        }
    })


def _match_clients_from_text(text_lines, umbral=0.52):
    """
    Para cada línea OCR busca el cliente cuyo nombre sea más similar
    usando difflib.SequenceMatcher.
    Solo acepta coincidencias con ratio >= umbral.
    Retorna tupla (clientes_encontrados, lineas_no_encontradas):
      - clientes_encontrados: lista de Clientes encontrados (sin duplicados)
      - lineas_no_encontradas: lista de strings de líneas que no tuvieron match
    """
    # Cargar todos los clientes una sola vez
    todos = list(Cliente.objects.values('id', 'nombre'))
    nombres_norm = [(c['id'], _normalizar(c['nombre'])) for c in todos]

    found = []
    no_encontradas = []
    seen_ids = set()

    for linea in text_lines:
        linea_norm = _normalizar(linea)
        # Ignorar líneas muy cortas o que parezcan números/códigos
        if len(linea_norm) < 5:
            continue

        mejor_ratio = 0.0
        mejor_id = None

        for cid, nombre_norm in nombres_norm:
            ratio = SequenceMatcher(None, linea_norm, nombre_norm).ratio()
            if ratio > mejor_ratio:
                mejor_ratio = ratio
                mejor_id = cid

        if mejor_ratio >= umbral and mejor_id not in seen_ids:
            try:
                cliente = Cliente.objects.get(pk=mejor_id)
                found.append(cliente)
                seen_ids.add(mejor_id)
            except Cliente.DoesNotExist:
                no_encontradas.append(linea.strip())
        else:
            # No se encontró coincidencia suficiente → línea sin registrar
            if mejor_ratio < umbral:
                no_encontradas.append(linea.strip())

    return found, no_encontradas


# ── Views ──────────────────────────────────────────────────────────────────────

class RutaListView(LoginRequiredMixin, ListView):
    model = RutaDia
    template_name = 'routes/list.html'
    context_object_name = 'rutas'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('conductor', 'peoneta', 'empresa')
        fecha = self.request.GET.get('fecha', '')
        conductor = self.request.GET.get('conductor', '')
        empresa = self.request.GET.get('empresa', '')
        modalidad = self.request.GET.get('modalidad', '').strip()
        if fecha:
            qs = qs.filter(fecha=fecha)
        if conductor:
            qs = qs.filter(conductor_id=conductor)
        if empresa:
            qs = qs.filter(empresa_id=empresa)
        if modalidad in {'estandar', 'falabella'}:
            qs = qs.filter(modalidad=modalidad)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['empresas'] = Empresa.objects.filter(activa=True).order_by('nombre')
        ctx['modalidad_actual'] = self.request.GET.get('modalidad', '').strip()
        return ctx

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except (OperationalError, ProgrammingError) as exc:
            text = str(exc).lower()
            _write_routes_log('routes_list_error.log', traceback.format_exc())
            if (
                'peoneta_id' in text
                or 'routes_rutadia' in text
                or 'rendiciones_rendicionreparto' in text
                or ("doesn't exist" in text and 'rendiciones_' in text)
            ):
                messages.error(
                    request,
                    'La base de datos del hosting esta desactualizada (faltan migraciones). '
                    'Ejecuta: python manage.py migrate',
                )
                return redirect('dashboard:index')
            raise
        except Exception:
            _write_routes_log('routes_list_error.log', traceback.format_exc())
            messages.error(
                request,
                'Ocurrio un error al abrir Ruta del dia. Revisa tmp/routes_list_error.log para el detalle.',
            )
            return redirect('dashboard:index')


class RutaCreateView(LoginRequiredMixin, CreateView):
    model = RutaDia
    form_class = RutaDiaForm
    template_name = 'routes/form.html'
    success_url = reverse_lazy('routes:list')

    def form_valid(self, form):
        messages.success(self.request, 'Ruta del día creada exitosamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nueva Ruta del Día'
        return ctx


class RutaUpdateView(LoginRequiredMixin, UpdateView):
    model = RutaDia
    form_class = RutaDiaForm
    template_name = 'routes/form.html'
    success_url = reverse_lazy('routes:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Editar Ruta'
        return ctx


class RutaDeleteView(LoginRequiredMixin, DeleteView):
    model = RutaDia
    template_name = 'routes/confirm_delete.html'
    success_url = reverse_lazy('routes:list')


class RutaDetailView(LoginRequiredMixin, DetailView):
    model = RutaDia
    template_name = 'routes/detail.html'
    context_object_name = 'ruta'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['paradas'] = self.object.paradas.select_related(
            'entrega__cliente',
            'entrega__empresa',
            'entrega__conductor',
        ).prefetch_related('entrega__pagos').order_by('orden')
        ctx['rendicion'] = getattr(self.object, 'rendicion', None)
        return ctx


@login_required
def ruta_falabella_import(request):
    token = request.POST.get('upload_token') or request.GET.get('token') or ''
    payload = _falabella_payload_from_session(request, token) if token else None
    empresas_choices = payload.get('empresas', []) if payload else []
    patentes_choices = payload.get('patentes', []) if payload else []

    initial_data = {'action': 'upload'}
    if payload:
        form_data = payload.get('form_data') or {}
        initial_data.update(
            {
                'fecha': form_data.get('fecha') or '',
                'conductor': form_data.get('conductor') or '',
                'peoneta': form_data.get('peoneta') or '',
                'upload_token': token,
                'action': 'create',
            }
        )

    if request.method == 'POST':
        form = RutaFalabellaExcelForm(
            request.POST,
            request.FILES,
            empresas_choices=empresas_choices,
            patentes_choices=patentes_choices,
        )
        if form.is_valid():
            action = (form.cleaned_data.get('action') or '').strip().lower()

            if action == 'upload':
                try:
                    rows, empresas, patentes = _read_excel_rows(form.cleaned_data['archivo_excel'])
                except ValueError as exc:
                    form.add_error('archivo_excel', str(exc))
                else:
                    token = uuid.uuid4().hex
                    request.session[f'{FALABELLA_SESSION_PREFIX}{token}'] = {
                        'rows': rows,
                        'empresas': empresas,
                        'patentes': patentes,
                        'form_data': {
                            'fecha': form.cleaned_data['fecha'].isoformat() if form.cleaned_data.get('fecha') else '',
                            'conductor': form.cleaned_data['conductor'].pk if form.cleaned_data.get('conductor') else '',
                            'peoneta': form.cleaned_data['peoneta'].pk if form.cleaned_data.get('peoneta') else '',
                        },
                    }
                    request.session.modified = True
                    messages.success(request, f'Se cargaron {len(rows)} filas del Excel. Selecciona empresa y patente para crear la ruta.')
                    return redirect(f"{reverse_lazy('routes:falabella_import')}?token={token}")

            if action == 'create':
                token = form.cleaned_data.get('upload_token')
                payload = _falabella_payload_from_session(request, token)
                if not payload:
                    form.add_error(None, 'La sesión de carga expiró. Vuelve a subir el archivo.')
                else:
                    empresa_name = form.cleaned_data['empresa_archivo']
                    patente_name = form.cleaned_data['patente_archivo']
                    rows = [
                        row for row in payload.get('rows', [])
                        if row.get('empresa') == empresa_name and row.get('patente') == patente_name
                    ]
                    if not rows:
                        form.add_error(None, 'No hay filas para la combinación empresa/patente seleccionada.')
                    else:
                        empresa = form.cleaned_data.get('empresa_objetivo')
                        if empresa is None:
                            empresa = Empresa.objects.filter(nombre__iexact=empresa_name).first()

                        reutilizada = False
                        try:
                            with transaction.atomic():
                                ruta_existente = RutaDia.objects.select_for_update().filter(
                                    fecha=form.cleaned_data['fecha'],
                                    conductor=form.cleaned_data['conductor'],
                                ).first()

                                if ruta_existente and ruta_existente.modalidad != 'falabella':
                                    form.add_error(
                                        None,
                                        'El conductor ya tiene una ruta no Falabella para esa fecha. Selecciona otra fecha o conductor.',
                                    )
                                    raise ValueError('Ruta existente no compatible')

                                if ruta_existente:
                                    ruta = ruta_existente
                                    ruta.empresa = empresa
                                    ruta.peoneta = form.cleaned_data['peoneta']
                                    ruta.modalidad = 'falabella'
                                    ruta.patente = patente_name
                                    ruta.estado = 'planificada'
                                    ruta.observacion = f'Ruta Falabella importada desde Excel ({empresa_name}/{patente_name}).'
                                    ruta.save(
                                        update_fields=[
                                            'empresa',
                                            'peoneta',
                                            'modalidad',
                                            'patente',
                                            'estado',
                                            'observacion',
                                            'fecha_actualizacion',
                                        ]
                                    )
                                    ruta.paradas.all().delete()
                                    reutilizada = True
                                else:
                                    ruta = RutaDia.objects.create(
                                        fecha=form.cleaned_data['fecha'],
                                        empresa=empresa,
                                        conductor=form.cleaned_data['conductor'],
                                        peoneta=form.cleaned_data['peoneta'],
                                        modalidad='falabella',
                                        patente=patente_name,
                                        estado='planificada',
                                        observacion=f'Ruta Falabella importada desde Excel ({empresa_name}/{patente_name}).',
                                    )

                                for idx, row in enumerate(rows, start=1):
                                    direccion = row.get('direccion', '').strip()
                                    localidad = row.get('localidad', '').strip()[:100]
                                    cliente = Cliente.objects.filter(
                                        empresa=empresa,
                                        direccion__iexact=direccion,
                                        comuna__iexact=localidad,
                                    ).first()
                                    if cliente is None:
                                        nombre_sugerido = f'Entrega Falabella {idx}'
                                        cliente = Cliente.objects.create(
                                            nombre=nombre_sugerido[:200],
                                            empresa=empresa,
                                            comuna=localidad,
                                            direccion=direccion,
                                            ciudad=localidad,
                                            region="Region de O'Higgins",
                                        )
                                    updates = []
                                    if empresa and cliente.empresa_id != empresa.pk:
                                        cliente.empresa = empresa
                                        updates.append('empresa')
                                    if direccion and not cliente.direccion:
                                        cliente.direccion = direccion
                                        updates.append('direccion')
                                    if localidad and not cliente.comuna:
                                        cliente.comuna = localidad
                                        updates.append('comuna')
                                    if updates:
                                        updates.append('fecha_actualizacion')
                                        cliente.save(update_fields=updates)

                                    entrega = Entrega.objects.create(
                                        cliente=cliente,
                                        empresa=empresa,
                                        conductor=ruta.conductor,
                                        estado='pendiente',
                                        fecha_programada=ruta.fecha,
                                        descripcion='Entrega Falabella',
                                    )

                                    # Evita choques de unicidad (ruta, orden) cuando la ruta ya tiene paradas previas.
                                    parada, creada_parada = ParadaRuta.objects.update_or_create(
                                        ruta=ruta,
                                        orden=idx,
                                        defaults={'entrega': entrega},
                                    )
                                    if not creada_parada:
                                        ParadaUbicacionCandidata.objects.filter(parada=parada).delete()
                                        ParadaFalabellaMeta.objects.filter(parada=parada).delete()

                                    ParadaUbicacionCandidata.objects.filter(parada=parada).delete()
                                    ParadaFalabellaMeta.objects.update_or_create(
                                        parada=parada,
                                        defaults={
                                            'direccion_original': row.get('direccion', ''),
                                            'localidad_original': row.get('localidad', ''),
                                            'contacto_original': '',
                                            'estado_direccion': 'pendiente_busqueda',
                                        },
                                    )

                                    geocode_coords = _geocode_free_address(row.get('direccion', ''), row.get('localidad', ''))
                                    if geocode_coords:
                                        cliente.latitud, cliente.longitud = geocode_coords
                                        cliente.save(update_fields=['latitud', 'longitud', 'fecha_actualizacion'])
                                        ParadaFalabellaMeta.objects.filter(parada=parada).update(estado_direccion='confirmada_manual')
                                    else:
                                        # Mantener coordenadas existentes si ya estaban confirmadas para no perder el pin.
                                        if cliente.latitud is None or cliente.longitud is None:
                                            ParadaFalabellaMeta.objects.filter(parada=parada).update(estado_direccion='requiere_llamada_cliente')
                                        else:
                                            ParadaFalabellaMeta.objects.filter(parada=parada).update(estado_direccion='confirmada_manual')

                                if reutilizada:
                                    ruta.paradas.filter(orden__gt=len(rows)).delete()

                                _falabella_optimizar_con_anclas(ruta)
                        except Exception as exc:
                            if not form.non_field_errors():
                                form.add_error(None, f'No fue posible crear la ruta: {exc}')
                        else:
                            accion = 'actualizada' if reutilizada else 'creada'
                            messages.success(
                                request,
                                f'Ruta Falabella {accion} con {len(rows)} paradas. Navegación basada en direcciones de Google Maps.',
                            )
                            return redirect('routes:falabella_detail', pk=ruta.pk)
    else:
        form = RutaFalabellaExcelForm(
            initial=initial_data,
            empresas_choices=empresas_choices,
            patentes_choices=patentes_choices,
        )

    preview_rows = []
    if payload:
        preview_rows = payload.get('rows', [])[:20]
    return render(
        request,
        'routes/falabella_import.html',
        {
            'form': form,
            'preview_rows': preview_rows,
            'has_payload': bool(payload),
            'titulo': 'Nueva Ruta Falabella (Excel)',
        },
    )


@login_required
def ruta_falabella_detail(request, pk):
    ruta = get_object_or_404(
        RutaDia.objects.select_related('empresa', 'conductor', 'peoneta'),
        pk=pk,
        modalidad='falabella',
    )
    paradas = ruta.paradas.select_related(
        'entrega__cliente',
        'falabella_meta',
    ).prefetch_related(
        'ubicaciones_candidatas',
    ).order_by('orden')
    return render(
        request,
        'routes/falabella_detail.html',
        {
            'ruta': ruta,
            'paradas': paradas,
            'coord_rengo': COORD_RANGO_RENGO,
            'coord_rancagua': COORD_RANGO_RANCAGUA,
        },
    )


@login_required
def falabella_seleccionar_candidato(request, pk, parada_id, candidato_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk, modalidad='falabella')
    parada = get_object_or_404(ParadaRuta, pk=parada_id, ruta=ruta)
    candidato = get_object_or_404(ParadaUbicacionCandidata, pk=candidato_id, parada=parada)

    try:
        ParadaUbicacionCandidata.objects.filter(parada=parada).update(seleccionada=False)
        candidato.seleccionada = True
        candidato.save(update_fields=['seleccionada'])

        cliente = parada.entrega.cliente
        cliente.latitud = candidato.latitud
        cliente.longitud = candidato.longitud
        cliente.save(update_fields=['latitud', 'longitud', 'fecha_actualizacion'])

        meta = getattr(parada, 'falabella_meta', None)
        if meta:
            meta.estado_direccion = 'confirmada_manual'
            meta.save(update_fields=['estado_direccion', 'fecha_actualizacion'])

        _falabella_optimizar_con_anclas(ruta)
    except Exception as exc:
        return JsonResponse({'error': f'No fue posible actualizar la ubicación: {exc}'}, status=400)

    return JsonResponse({'ok': True, 'cliente': cliente.nombre})


@login_required
def falabella_actualizar_direccion(request, pk, parada_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk, modalidad='falabella')
    parada = get_object_or_404(ParadaRuta.objects.select_related('entrega__cliente'), pk=parada_id, ruta=ruta)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'error': 'Solicitud inválida.'}, status=400)

    direccion = str(payload.get('direccion', '') or '').strip()
    localidad = str(payload.get('localidad', '') or '').strip()
    if not direccion:
        return JsonResponse({'error': 'Debes ingresar una dirección.'}, status=400)

    cliente = parada.entrega.cliente
    cliente.direccion = direccion
    updates = ['direccion', 'fecha_actualizacion']
    if localidad:
        cliente.comuna = localidad[:100]
        updates.append('comuna')
        if not cliente.ciudad:
            cliente.ciudad = localidad[:100]
            updates.append('ciudad')

    geocode_coords = _geocode_free_address(direccion, localidad)
    if geocode_coords:
        cliente.latitud, cliente.longitud = geocode_coords
        updates.extend(['latitud', 'longitud'])
    else:
        # Si la nueva dirección no coincide con alta confianza, no borrar coordenadas existentes.
        if cliente.latitud is None or cliente.longitud is None:
            cliente.latitud = None
            cliente.longitud = None
            updates.extend(['latitud', 'longitud'])

    cliente.save(update_fields=list(dict.fromkeys(updates)))

    ParadaUbicacionCandidata.objects.filter(parada=parada).delete()
    ParadaFalabellaMeta.objects.update_or_create(
        parada=parada,
        defaults={
            'direccion_original': direccion,
            'localidad_original': localidad,
            'contacto_original': getattr(getattr(parada, 'falabella_meta', None), 'contacto_original', ''),
            'estado_direccion': 'confirmada_manual' if (geocode_coords or (cliente.latitud is not None and cliente.longitud is not None)) else 'requiere_llamada_cliente',
        },
    )
    cliente.refresh_from_db()
    _falabella_optimizar_con_anclas(ruta)

    estado = 'confirmada_manual' if (geocode_coords or (cliente.latitud is not None and cliente.longitud is not None)) else 'requiere_llamada_cliente'
    return JsonResponse(
        {
            'ok': True,
            'estado': estado,
            'candidatos': 0,
            'geocodificada_auto': bool(geocode_coords),
            'cliente': cliente.nombre,
            'latitud': float(cliente.latitud) if cliente.latitud is not None else None,
            'longitud': float(cliente.longitud) if cliente.longitud is not None else None,
            'google_maps_busqueda': cliente.url_google_maps_busqueda,
            'google_maps_coordenadas': cliente.url_google_maps if cliente.tiene_coordenadas() else '',
            'waze': cliente.url_waze if cliente.tiene_coordenadas() else '',
        }
    )


@login_required
def falabella_marcar_entregado(request, pk, parada_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk, modalidad='falabella')
    parada = get_object_or_404(ParadaRuta.objects.select_related('entrega'), pk=parada_id, ruta=ruta)
    entrega = parada.entrega
    if entrega.estado == 'entregado':
        entrega.estado = 'pendiente'
        entrega.fecha_entrega = None
    else:
        entrega.estado = 'entregado'
        entrega.fecha_entrega = timezone.now()
    entrega.save(update_fields=['estado', 'fecha_entrega', 'fecha_actualizacion'])

    return JsonResponse(
        {
            'ok': True,
            'estado': entrega.estado,
            'estado_display': entrega.get_estado_display(),
            'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
        }
    )


@login_required
def falabella_reoptimizar(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    ruta = get_object_or_404(RutaDia, pk=pk, modalidad='falabella')
    con_coords, pendientes = _falabella_optimizar_con_anclas(ruta)
    return JsonResponse({'ok': True, 'con_coords': con_coords, 'pendientes': pendientes})


@login_required
def procesar_foto_ruta(request, pk):
    """
    Procesa la foto y devuelve facturas detectadas con sugerencia de clientes.
    """
    ruta = get_object_or_404(RutaDia, pk=pk)
    if not ruta.foto_hoja_ruta:
        return JsonResponse({'error': 'No hay foto subida.'}, status=400)

    facturas_raw, raw_text = _ocr_extract_facturas(ruta.foto_hoja_ruta.path)

    if not facturas_raw:
        # Fallback compatible: usa texto libre para detectar clientes al menos por nombre.
        # Evita que el flujo falle completo cuando Gemini responde fuera del formato esperado.
        lineas_texto = [
            l.strip() for l in re.split(r'[\n\r]+', raw_text or '') if l.strip()
        ]
        clientes_encontrados, no_encontrados = _match_clients_from_text(lineas_texto, umbral=0.60)
        texto_completo = (raw_text or '\n'.join(lineas_texto)).strip()
        ruta.texto_extraido = texto_completo
        ruta.ocr_facturas_raw = []
        ruta.save(update_fields=['texto_extraido', 'ocr_facturas_raw'])

        return JsonResponse({
            'texto_extraido': texto_completo,
            'facturas': [],
            'clientes': [
                {'id': c.pk, 'nombre': c.nombre, 'comuna': c.comuna}
                for c in clientes_encontrados
            ],
            'no_encontrados': no_encontrados,
            'warning': 'No se pudo estructurar por factura automáticamente. Se aplicó detección por nombre para que puedas continuar.',
        })

    texto_completo = '\n'.join(
        f"{row.get('documento', '')} | {row.get('nombre_cliente', '')} | {row.get('direccion_cliente', '')}"
        for row in facturas_raw
    )
    facturas_detectadas, clientes_encontrados, no_encontrados = _match_clients_from_facturas(facturas_raw)

    # DEBUG: Log los totales extraídos por OCR
    for fac in facturas_detectadas:
        _write_routes_log('ocr_debug.log', f'Factura {fac.get("documento")} - Total: {fac.get("total")} (type: {type(fac.get("total")).__name__})')

    ruta.texto_extraido = texto_completo
    ruta.ocr_facturas_raw = facturas_detectadas
    ruta.save(update_fields=['texto_extraido', 'ocr_facturas_raw'])

    data = {
        'texto_extraido': texto_completo,
        'facturas': facturas_detectadas,
        'clientes': clientes_encontrados,
        'no_encontrados': no_encontrados,
    }
    return JsonResponse(data)


@login_required
def buscar_por_texto_ruta(request, pk):
    """
    Recibe un texto plano (un nombre por línea) vía POST,
    corre el mismo matching que el OCR y devuelve los clientes encontrados.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    ruta = get_object_or_404(RutaDia, pk=pk)
    try:
        body = json.loads(request.body)
        texto = body.get('texto', '')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    if not lineas:
        return JsonResponse({'error': 'El texto está vacío.'}, status=400)

    # Para texto libre usamos un umbral de confianza más alto (0.72):
    # el usuario escribe nombres limpios y sólo deben considerarse "encontrados"
    # los clientes con una coincidencia muy cercana. Así, nombres que no existen
    # en la BD aparecerán siempre en la sección "no encontrados" con botón Registrar.
    clientes_encontrados, no_encontrados = _match_clients_from_text(lineas, umbral=0.72)
    return JsonResponse({
        'clientes': [
            {'id': c.pk, 'nombre': c.nombre, 'comuna': c.comuna}
            for c in clientes_encontrados
        ],
        'no_encontrados': no_encontrados,
    })


@login_required
def buscar_clientes(request):
    """
    Búsqueda JSON de clientes por nombre o comuna (para agregar manualmente).
    """
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'clientes': []})
    clientes = Cliente.objects.filter(
        Q(nombre__icontains=q) | Q(comuna__icontains=q)
    )[:20]
    return JsonResponse({'clientes': [
        {'id': c.pk, 'nombre': c.nombre, 'comuna': c.comuna}
        for c in clientes
    ]})


@login_required
def crear_entregas_desde_ruta(request, pk):
    """
    Crea entregas para la ruta y reoptimiza paradas.

    Compatibilidad:
    - Flujo nuevo: recibe `lineas` con datos por factura OCR.
    - Flujo legado: recibe `cliente_ids`.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)

    try:
        body = json.loads(request.body)
        cliente_ids = body.get('cliente_ids', [])
        lineas = body.get('lineas', [])
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    if not cliente_ids and not lineas:
        return JsonResponse({'error': 'No se enviaron clientes.'}, status=400)

    entregas_creadas = []
    entregas_existentes = []
    pendientes = 0

    entrega_ids_existentes_en_ruta = list(
        ruta.paradas.values_list('entrega_id', flat=True)
    )

    if lineas:
        facturas_en_payload = []
        for linea in lineas:
            numero_factura = str((linea or {}).get('documento', '') or '').strip()
            if numero_factura:
                facturas_en_payload.append(numero_factura)
        if len(set(facturas_en_payload)) != len(facturas_en_payload):
            return JsonResponse({'error': 'Hay números de factura duplicados en la misma ruta.'}, status=400)

        for linea in lineas:
            numero_factura = str((linea or {}).get('documento', '') or '').strip()
            cliente_id = (linea or {}).get('cliente_id')
            nombre_ocr = str((linea or {}).get('nombre_cliente', '') or '').strip()
            direccion_ocr = str((linea or {}).get('direccion_cliente', '') or '').strip()
            comuna_ocr = str((linea or {}).get('comuna', '') or '').strip()
            total_factura = _safe_total((linea or {}).get('total'))
            dia_factura = _safe_int((linea or {}).get('dia'))
            mes_factura = _safe_int((linea or {}).get('mes'))
            cond_pago = str((linea or {}).get('cond_pago', '') or '').strip()

            cliente = None
            if cliente_id:
                cliente = Cliente.objects.filter(pk=cliente_id).first()

            requiere_revision = not bool(cliente)
            if requiere_revision:
                pendientes += 1

            defaults_linea = {
                'nombre_cliente_ocr': nombre_ocr,
                'direccion_ocr': direccion_ocr,
                'comuna_ocr': comuna_ocr,
                'dia_factura': dia_factura,
                'mes_factura': mes_factura,
                'total_factura': total_factura,
                'transportista': ruta.conductor.get_full_name() or ruta.conductor.username,
                'condicion_pago': cond_pago,
                'cliente': cliente,
                'estado_validacion': 'pendiente' if requiere_revision else 'confirmada',
                'requiere_revision': requiere_revision,
                'observacion_validacion': 'Cliente pendiente de selección' if requiere_revision else '',
            }

            lookup_numero = numero_factura or f"tmp-{uuid.uuid4().hex[:8]}"
            factura_linea, _ = FacturaOCRLinea.objects.update_or_create(
                ruta=ruta,
                numero_factura=lookup_numero,
                defaults=defaults_linea,
            )

            if requiere_revision:
                continue

            alias_key = _build_ocr_alias_key(nombre_ocr, direccion_ocr, comuna_ocr)
            if alias_key and alias_key != '||||':
                alias = ClienteOCRAlias.objects.filter(clave_normalizada=alias_key).first()
                if alias is None:
                    ClienteOCRAlias.objects.create(
                        clave_normalizada=alias_key,
                        nombre_ocr=nombre_ocr,
                        direccion_ocr=direccion_ocr,
                        comuna_ocr=comuna_ocr,
                        cliente=cliente,
                        audit_usuario=request.user,
                    )
                elif alias.cliente_id != cliente.pk:
                    alias.bloqueado_por_conflicto = True
                    alias.audit_usuario = request.user
                    alias.save(update_fields=['bloqueado_por_conflicto', 'audit_usuario', 'fecha_actualizacion'])
                else:
                    alias.bloqueado_por_conflicto = False
                    alias.audit_usuario = request.user
                    alias.nombre_ocr = nombre_ocr
                    alias.direccion_ocr = direccion_ocr
                    alias.comuna_ocr = comuna_ocr
                    alias.save(update_fields=[
                        'bloqueado_por_conflicto',
                        'audit_usuario',
                        'nombre_ocr',
                        'direccion_ocr',
                        'comuna_ocr',
                        'fecha_actualizacion',
                    ])

            # Actualiza ficha de cliente desde OCR validado
            updates_cliente = []
            if direccion_ocr and cliente.direccion != direccion_ocr:
                cliente.direccion = direccion_ocr
                updates_cliente.append('direccion')
            if comuna_ocr and cliente.comuna != comuna_ocr:
                cliente.comuna = comuna_ocr
                updates_cliente.append('comuna')
            if updates_cliente:
                updates_cliente.append('fecha_actualizacion')
                cliente.save(update_fields=updates_cliente)

            if numero_factura:
                entrega, created = Entrega.objects.get_or_create(
                    cliente=cliente,
                    fecha_programada=ruta.fecha,
                    conductor=ruta.conductor,
                    numero_factura_ref=numero_factura,
                    defaults={
                        'estado': 'pendiente',
                        'empresa': ruta.empresa,
                    }
                )
            else:
                entrega = Entrega.objects.create(
                    cliente=cliente,
                    fecha_programada=ruta.fecha,
                    conductor=ruta.conductor,
                    estado='pendiente',
                    empresa=ruta.empresa,
                )
                created = True

            entrega.total_factura_ref = total_factura
            entrega.dia_factura_ref = dia_factura
            entrega.mes_factura_ref = mes_factura
            entrega.condicion_pago_ref = cond_pago
            entrega.direccion_factura_ref = direccion_ocr
            entrega.comuna_factura_ref = comuna_ocr
            entrega.save(update_fields=[
                'total_factura_ref',
                'dia_factura_ref',
                'mes_factura_ref',
                'condicion_pago_ref',
                'direccion_factura_ref',
                'comuna_factura_ref',
                'fecha_actualizacion',
            ])

            factura_linea.entrega = entrega
            factura_linea.cliente = cliente
            factura_linea.audit_usuario = request.user
            factura_linea.fecha_validacion = timezone.now()
            factura_linea.estado_validacion = 'confirmada'
            factura_linea.requiere_revision = False
            factura_linea.observacion_validacion = ''
            factura_linea.save(update_fields=[
                'entrega',
                'cliente',
                'audit_usuario',
                'fecha_validacion',
                'estado_validacion',
                'requiere_revision',
                'observacion_validacion',
                'fecha_actualizacion',
            ])

            if created:
                entregas_creadas.append(entrega)
            else:
                entregas_existentes.append(entrega)
    else:
        clientes = Cliente.objects.filter(pk__in=cliente_ids)
        for cliente in clientes:
            entrega, created = Entrega.objects.get_or_create(
                cliente=cliente,
                fecha_programada=ruta.fecha,
                conductor=ruta.conductor,
                defaults={
                    'estado': 'pendiente',
                    'empresa': ruta.empresa,
                }
            )
            if created:
                entregas_creadas.append(entrega)
            else:
                entregas_existentes.append(entrega)

    entregas_previas = list(Entrega.objects.filter(pk__in=entrega_ids_existentes_en_ruta))
    todas_las_entregas = entregas_previas + entregas_creadas + entregas_existentes

    # Evitar duplicados manteniendo orden estable
    entregas_unicas = []
    seen_ids = set()
    for entrega in todas_las_entregas:
        if entrega.pk in seen_ids:
            continue
        seen_ids.add(entrega.pk)
        entregas_unicas.append(entrega)
    todas_las_entregas = entregas_unicas

    # Optimizar con Nearest Neighbor
    points = []
    for e in todas_las_entregas:
        c = e.cliente
        if c.latitud is not None and c.longitud is not None:
            points.append({
                'id': e.pk,
                'lat': float(c.latitud),
                'lon': float(c.longitud),
                'nombre': c.nombre,
            })

    if len(points) >= 2:
        optimized_ids = _nearest_neighbor_route(points)
        optimized_set = set(optimized_ids)
        # Mantiene en la ruta entregas sin coordenadas (o fuera del optimizador)
        # para que agregar cliente manual nunca "desaparezca" del resultado.
        restantes_ids = [e.pk for e in todas_las_entregas if e.pk not in optimized_set]
        ordered_ids = optimized_ids + restantes_ids
    else:
        ordered_ids = [e.pk for e in todas_las_entregas]

    # Guardar paradas: solo reemplaza las de esta ruta
    ruta.paradas.all().delete()
    paradas_guardadas = []
    for i, eid in enumerate(ordered_ids, start=1):
        try:
            entrega = Entrega.objects.get(pk=eid)
            parada = ParadaRuta.objects.create(ruta=ruta, entrega=entrega, orden=i)
            c = entrega.cliente
            resumen_pago = _resumen_pago_entrega(entrega)
            paradas_guardadas.append({
                'orden': i,
                'entrega_id': eid,
                'cliente_id': c.pk,
                'cliente': c.nombre,
                'comuna': c.comuna,
                'observaciones': c.observaciones or '',
                'empresa': entrega.empresa.nombre if entrega.empresa else '',
                'conductor': entrega.conductor.get_full_name() if entrega.conductor else '',
                'descripcion': entrega.descripcion or '',
                'observacion_entrega': entrega.observacion or '',
                'fecha_programada': entrega.fecha_programada.strftime('%d-%m-%Y') if entrega.fecha_programada else '',
                'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
                'numero_factura': entrega.numero_factura_ref,
                'total_factura_ref': float(entrega.total_factura_ref or 0),
                'equivalencia_entregas': int(c.equivalencia_entregas or 1),
                'estado': entrega.estado,
                'estado_display': entrega.get_estado_display(),
                'estado_pago': resumen_pago['estado_pago'],
                'estado_pago_display': resumen_pago['estado_pago_display'],
                'total_pagado': resumen_pago['total_pagado'],
                'lat': float(c.latitud) if c.latitud is not None else None,
                'lon': float(c.longitud) if c.longitud is not None else None,
            })
        except Entrega.DoesNotExist:
            pass

    return JsonResponse({
        'creadas': len(entregas_creadas),
        'existentes': len(entregas_existentes),
        'pendientes': pendientes,
        'total_paradas': len(paradas_guardadas),
        'paradas': paradas_guardadas,
    })


@login_required
def optimizar_ruta(request, pk):
    """
    Re-optimiza las paradas existentes de una ruta ya creada.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)
    paradas_qs = ruta.paradas.select_related('entrega__cliente').all()

    points = []
    for p in paradas_qs:
        c = p.entrega.cliente
        if c.latitud and c.longitud:
            points.append({
                'id': p.entrega.pk,
                'lat': float(c.latitud),
                'lon': float(c.longitud),
                'nombre': c.nombre,
            })

    if len(points) < 2:
        return JsonResponse({'error': 'Se necesitan al menos 2 paradas con coordenadas.'}, status=400)

    ordered_ids = _nearest_neighbor_route(points)

    ruta.paradas.all().delete()
    paradas_guardadas = []
    for i, eid in enumerate(ordered_ids, start=1):
        entrega = Entrega.objects.get(pk=eid)
        ParadaRuta.objects.create(ruta=ruta, entrega=entrega, orden=i)
        c = entrega.cliente
        resumen_pago = _resumen_pago_entrega(entrega)
        paradas_guardadas.append({
            'orden': i,
            'entrega_id': eid,
            'cliente_id': c.pk,
            'cliente': c.nombre,
            'comuna': c.comuna or '–',
            'observaciones': c.observaciones or '',
            'empresa': entrega.empresa.nombre if entrega.empresa else '',
            'conductor': entrega.conductor.get_full_name() if entrega.conductor else '',
            'descripcion': entrega.descripcion or '',
            'observacion_entrega': entrega.observacion or '',
            'fecha_programada': entrega.fecha_programada.strftime('%d-%m-%Y') if entrega.fecha_programada else '',
            'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
            'estado': entrega.estado,
            'estado_display': entrega.get_estado_display(),
            'estado_pago': resumen_pago['estado_pago'],
            'estado_pago_display': resumen_pago['estado_pago_display'],
            'total_pagado': resumen_pago['total_pagado'],
            'lat': float(c.latitud) if c.latitud else None,
            'lon': float(c.longitud) if c.longitud else None,
        })

    return JsonResponse({'ruta_optimizada': paradas_guardadas, 'total': len(paradas_guardadas)})


# ── Navegación en Vivo ────────────────────────────────────────────────────────

@login_required
def navegacion_ruta(request, pk):
    """
    Vista de navegación en vivo para el conductor.
    Muestra el mapa con todas las paradas y permite re-optimizar desde la
    ubicación actual del dispositivo.
    """
    ruta = get_object_or_404(RutaDia, pk=pk)
    paradas = ruta.paradas.select_related('entrega__cliente').order_by('orden')
    ctx = {
        'ruta': ruta,
        'paradas': paradas,
    }
    return render(request, 'routes/navegacion.html', ctx)


@login_required
def reoptimizar_desde_posicion(request, pk):
    """
    Re-optimiza las paradas de una ruta usando la posición actual del conductor
    como punto de partida del algoritmo Nearest Neighbor.
    Recibe POST JSON: {lat: float, lon: float}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)

    try:
        body = json.loads(request.body)
        lat_actual = float(body.get('lat', 0))
        lon_actual = float(body.get('lon', 0))
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Datos de posición inválidos.'}, status=400)

    paradas_qs = ruta.paradas.select_related('entrega__cliente').all()
    if not paradas_qs.exists():
        return JsonResponse({'error': 'La ruta no tiene paradas.'}, status=400)

    # Punto origen: posición actual del conductor
    points = [{'id': 'origen', 'lat': lat_actual, 'lon': lon_actual}]
    for p in paradas_qs:
        c = p.entrega.cliente
        if c.latitud and c.longitud:
            points.append({
                'id': p.entrega.pk,
                'lat': float(c.latitud),
                'lon': float(c.longitud),
                'nombre': c.nombre,
                'estado': p.entrega.estado,
            })

    # Solo re-optimizar las que no están entregadas
    pendientes = [p for p in points[1:] if p.get('estado') != 'entregado']
    entregadas = [p for p in points[1:] if p.get('estado') == 'entregado']

    if len(pendientes) >= 2:
        # Re-run Nearest Neighbor desde posición actual
        all_points = [points[0]] + pendientes
        ordered_ids = _nearest_neighbor_route(all_points)[1:]  # quitar origen
    elif len(pendientes) == 1:
        ordered_ids = [pendientes[0]['id']]
    else:
        return JsonResponse({'mensaje': 'Todas las entregas están completadas.', 'paradas': []})

    # Reordenar paradas en BD
    ruta.paradas.all().delete()
    orden = 1
    nueva_lista = []

    for eid in ordered_ids:
        try:
            entrega = Entrega.objects.get(pk=eid)
            ParadaRuta.objects.create(ruta=ruta, entrega=entrega, orden=orden)
            c = entrega.cliente
            resumen_pago = _resumen_pago_entrega(entrega)
            nueva_lista.append({
                'orden': orden,
                'entrega_id': eid,
                'cliente_id': c.pk,
                'cliente': c.nombre,
                'comuna': c.comuna or '–',
                'observaciones': c.observaciones or '',
                'empresa': entrega.empresa.nombre if entrega.empresa else '',
                'conductor': entrega.conductor.get_full_name() if entrega.conductor else '',
                'descripcion': entrega.descripcion or '',
                'observacion_entrega': entrega.observacion or '',
                'fecha_programada': entrega.fecha_programada.strftime('%d-%m-%Y') if entrega.fecha_programada else '',
                'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
                'estado': entrega.estado,
                'estado_display': entrega.get_estado_display(),
                'estado_pago': resumen_pago['estado_pago'],
                'estado_pago_display': resumen_pago['estado_pago_display'],
                'total_pagado': resumen_pago['total_pagado'],
                'lat': float(c.latitud) if c.latitud else None,
                'lon': float(c.longitud) if c.longitud else None,
                'waze': c.url_waze if c.tiene_coordenadas else None,
                'gmaps': c.url_google_maps if c.tiene_coordenadas else None,
            })
            orden += 1
        except Entrega.DoesNotExist:
            pass

    # Re-agregar las ya entregadas al final (sin cambiar BD)
    for p_data in entregadas:
        try:
            entrega = Entrega.objects.get(pk=p_data['id'])
            ParadaRuta.objects.create(ruta=ruta, entrega=entrega, orden=orden)
            c = entrega.cliente
            resumen_pago = _resumen_pago_entrega(entrega)
            nueva_lista.append({
                'orden': orden,
                'entrega_id': p_data['id'],
                'cliente_id': c.pk,
                'cliente': c.nombre,
                'comuna': c.comuna or '–',
                'observaciones': c.observaciones or '',
                'empresa': entrega.empresa.nombre if entrega.empresa else '',
                'conductor': entrega.conductor.get_full_name() if entrega.conductor else '',
                'descripcion': entrega.descripcion or '',
                'observacion_entrega': entrega.observacion or '',
                'fecha_programada': entrega.fecha_programada.strftime('%d-%m-%Y') if entrega.fecha_programada else '',
                'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
                'estado': entrega.estado,
                'estado_display': entrega.get_estado_display(),
                'estado_pago': resumen_pago['estado_pago'],
                'estado_pago_display': resumen_pago['estado_pago_display'],
                'total_pagado': resumen_pago['total_pagado'],
                'lat': float(c.latitud) if c.latitud else None,
                'lon': float(c.longitud) if c.longitud else None,
                'waze': c.url_waze if c.tiene_coordenadas else None,
                'gmaps': c.url_google_maps if c.tiene_coordenadas else None,
            })
            orden += 1
        except Entrega.DoesNotExist:
            pass

    return JsonResponse({'paradas': nueva_lista, 'total': len(nueva_lista)})


@login_required
def actualizar_estado_parada(request, pk):
    """
    Actualiza el estado de una Entrega directamente desde la navegación.
    POST JSON: {entrega_id: int, estado: str}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)

    try:
        body = json.loads(request.body)
        entrega_id = int(body.get('entrega_id'))
        nuevo_estado = body.get('estado')
    except (json.JSONDecodeError, ValueError, TypeError):
        return JsonResponse({'error': 'Datos inválidos.'}, status=400)

    estados_validos = [e[0] for e in Entrega.ESTADO_CHOICES]
    if nuevo_estado not in estados_validos:
        return JsonResponse({'error': 'Estado no válido.'}, status=400)

    from django.utils import timezone
    get_object_or_404(ParadaRuta, ruta=ruta, entrega_id=entrega_id)
    entrega = get_object_or_404(Entrega, pk=entrega_id)
    entrega.estado = nuevo_estado
    if nuevo_estado == 'entregado':
        if not entrega.fecha_entrega:
            entrega.fecha_entrega = timezone.now()
    else:
        entrega.fecha_entrega = None
    entrega.save(update_fields=['estado', 'fecha_entrega'])

    return JsonResponse({
        'ok': True,
        'estado': entrega.estado,
        'estado_display': entrega.get_estado_display(),
        'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
    })


def _serializar_paradas_ruta(ruta):
    paradas = ruta.paradas.select_related('entrega__cliente', 'entrega__empresa', 'entrega__conductor').prefetch_related('entrega__pagos').order_by('orden', 'id')
    data = []
    for p in paradas:
        entrega = p.entrega
        c = entrega.cliente
        resumen_pago = _resumen_pago_entrega(entrega)
        data.append({
            'orden': p.orden,
            'entrega_id': entrega.pk,
            'cliente': c.nombre,
            'comuna': c.comuna or '–',
            'observaciones': c.observaciones or '',
            'empresa': entrega.empresa.nombre if entrega.empresa else '',
            'conductor': entrega.conductor.get_full_name() if entrega.conductor else '',
            'descripcion': entrega.descripcion or '',
            'observacion_entrega': entrega.observacion or '',
            'fecha_programada': entrega.fecha_programada.strftime('%d-%m-%Y') if entrega.fecha_programada else '',
            'fecha_entrega': entrega.fecha_entrega.strftime('%d-%m-%Y %H:%M') if entrega.fecha_entrega else '',
            'estado': entrega.estado,
            'estado_display': entrega.get_estado_display(),
            'estado_pago': resumen_pago['estado_pago'],
            'estado_pago_display': resumen_pago['estado_pago_display'],
            'total_pagado': resumen_pago['total_pagado'],
            'lat': float(c.latitud) if c.latitud is not None else None,
            'lon': float(c.longitud) if c.longitud is not None else None,
        })
    return data


@login_required
def eliminar_parada_ruta(request, pk):
    """
    Elimina una parada de la ruta (cliente) y reordena la secuencia.
    POST JSON: {entrega_id: int}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    ruta = get_object_or_404(RutaDia, pk=pk)

    try:
        body = json.loads(request.body)
        entrega_id = int(body.get('entrega_id'))
    except (json.JSONDecodeError, ValueError, TypeError):
        return JsonResponse({'error': 'Datos inválidos.'}, status=400)

    parada = ParadaRuta.objects.filter(ruta=ruta, entrega_id=entrega_id).first()
    if not parada:
        return JsonResponse({'error': 'La parada no existe en esta ruta.'}, status=404)

    parada.delete()

    # Reordena para evitar huecos en la numeración de paradas.
    paradas_restantes = list(ruta.paradas.order_by('orden', 'id'))
    for index, p in enumerate(paradas_restantes, start=1):
        if p.orden != index:
            p.orden = index
            p.save(update_fields=['orden'])

    paradas_data = _serializar_paradas_ruta(ruta)
    return JsonResponse({
        'ok': True,
        'total_paradas': len(paradas_data),
        'paradas': paradas_data,
    })
