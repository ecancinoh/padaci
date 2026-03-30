from io import BytesIO
from decimal import Decimal
from pathlib import Path
import traceback
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.db.utils import OperationalError, ProgrammingError
from django.http import HttpResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, DeleteView
from .forms import RendicionRepartoForm, build_formsets, get_clientes_ruta_nombres
from .models import RendicionReparto


def _build_autocompletado_desde_ruta(ruta):
    sugeridos = {
        'a': [],
        'b': [],
        'c': [],
        'd': [],
        'e': [],
    }
    resumen = {
        'total_items': 0,
        'paradas': 0,
        'entregas': 0,
        'pagos': 0,
    }

    if not ruta:
        return sugeridos, resumen

    from routes.models import ParadaRuta

    paradas = (
        ParadaRuta.objects.filter(ruta=ruta)
        .select_related('entrega__cliente')
        .prefetch_related('entrega__pagos')
        .order_by('orden', 'id')
    )

    seen = {clave: set() for clave in sugeridos.keys()}

    for parada in paradas:
        resumen['paradas'] += 1
        if not parada.entrega:
            continue

        entrega = parada.entrega
        cliente = entrega.cliente
        numero_ref = str(entrega.pk)
        resumen['entregas'] += 1

        if entrega.estado in {'fallido', 'devuelto'}:
            key = (numero_ref, Decimal('0'))
            if key not in seen['d']:
                sugeridos['d'].append({
                    'numero_factura': numero_ref,
                    'monto': Decimal('0'),
                })
                seen['d'].add(key)

        for pago in entrega.pagos.all():
            resumen['pagos'] += 1
            monto = pago.monto or Decimal('0')
            if monto <= 0:
                continue

            if pago.metodo == 'cheque':
                key = (numero_ref, cliente.nombre, monto)
                if key not in seen['a']:
                    sugeridos['a'].append({
                        'numero_factura': numero_ref,
                        'nombre_cliente': cliente.nombre,
                        'monto': monto,
                        'banco': '',
                    })
                    seen['a'].add(key)
                continue

            if pago.metodo == 'descuento':
                motivo = (pago.observacion or 'Descuento en pago')[:200]
                key = (numero_ref, motivo, monto)
                if key not in seen['b']:
                    sugeridos['b'].append({
                        'numero_factura': numero_ref,
                        'motivo': motivo,
                        'monto': monto,
                    })
                    seen['b'].add(key)
                continue

            if pago.metodo == 'credito':
                key = (numero_ref, cliente.nombre, monto)
                if key not in seen['c']:
                    sugeridos['c'].append({
                        'numero_factura': numero_ref,
                        'autoriza_credito': cliente.nombre,
                        'monto': monto,
                    })
                    seen['c'].add(key)
                continue

            if pago.metodo == 'transferencia':
                key = (numero_ref, monto)
                if key not in seen['e']:
                    sugeridos['e'].append({
                        'numero_factura': numero_ref,
                        'monto': monto,
                    })
                    seen['e'].add(key)
                continue

    resumen['total_items'] = sum(len(items) for items in sugeridos.values())
    return sugeridos, resumen


def _autocompletar_rendicion_desde_entregas(rendicion):
    """
    Regenera idempotentemente lineas A/B/C/E/D desde los pagos registrados por entrega en la ruta.
    Reglas de mapeo:
    - A: pago con metodo=cheque
    - B: pago con metodo=descuento
    - C: pago con metodo=credito
    - E: pago con metodo=transferencia
    - D: entrega con estado in {fallido, devuelto}
    
    Notas de idempotencia:
    - No borra items existentes (permite preservar ediciones manuales)
    - Solo crea items nuevos si no existen (validacion por numero_factura, cliente, monto)
    - Puede llamarse multiples veces sin crear duplicados
    """
    sugeridos, _ = _build_autocompletado_desde_ruta(rendicion.ruta)

    def _existe_credito_documento(nombre_cliente, monto):
        # Si ya existe una línea para ese cliente y monto, no sugerir otra (sin importar el número de factura)
        return rendicion.creditos_documentos.filter(
            nombre_cliente=nombre_cliente,
            monto=monto,
        ).exists()

    def _existe_credito_confianza(numero_factura, autoriza_credito, monto):
        return rendicion.creditos_confianza.filter(
            numero_factura=numero_factura,
            autoriza_credito=autoriza_credito,
            monto=monto,
        ).exists()

    def _existe_devolucion_parcial(numero_factura, motivo, monto):
        return rendicion.devoluciones_parciales.filter(
            numero_factura=numero_factura,
            motivo=motivo,
            monto=monto,
        ).exists()

    def _existe_deposito_transferencia(numero_factura, monto):
        return rendicion.depositos_transferencias.filter(
            numero_factura=numero_factura,
            monto=monto,
        ).exists()

    def _existe_factura_nula(numero_factura):
        return rendicion.facturas_nulas_detalle.filter(numero_factura=numero_factura).exists()

    for item in sugeridos['a']:
        if not _existe_credito_documento(item['nombre_cliente'], item['monto']):
            rendicion.creditos_documentos.create(**item)

    for item in sugeridos['b']:
        if not _existe_devolucion_parcial(item['numero_factura'], item['motivo'], item['monto']):
            rendicion.devoluciones_parciales.create(**item)

    for item in sugeridos['c']:
        if not _existe_credito_confianza(item['numero_factura'], item['autoriza_credito'], item['monto']):
            rendicion.creditos_confianza.create(**item)

    for item in sugeridos['d']:
        if not _existe_factura_nula(item['numero_factura']):
            rendicion.facturas_nulas_detalle.create(**item)

    for item in sugeridos['e']:
        if not _existe_deposito_transferencia(item['numero_factura'], item['monto']):
            rendicion.depositos_transferencias.create(**item)


def _write_rendiciones_log(filename, payload):
    base_dir = Path(__file__).resolve().parents[1]
    candidates = [
        base_dir / 'tmp' / filename,
        base_dir / filename,
        Path('/tmp') / filename,
    ]
    for path in candidates:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open('a', encoding='utf-8') as f:
                f.write(payload + '\n')
            return
        except Exception:
            continue


def _load_openpyxl():
    """Lazy import so app startup doesn't depend on openpyxl being installed."""
    from openpyxl import load_workbook

    return load_workbook


def _load_reportlab():
    """Lazy import so the app can boot even if reportlab is not installed."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    return letter, canvas


def _money(value):
    try:
        return f"$ {int(value):,}".replace(',', '.')
    except Exception:
        return '$ 0'


def _compact_rows(queryset, attrs):
    rows = []
    for obj in queryset:
        current = [getattr(obj, attr) for attr in attrs]
        if any(str(val).strip() for val in current):
            rows.append(current)
    return rows


class RendicionListView(LoginRequiredMixin, ListView):
    model = RendicionReparto
    template_name = 'rendiciones/list.html'
    context_object_name = 'rendiciones'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('ruta__conductor', 'ruta__empresa')
        fecha = self.request.GET.get('fecha', '').strip()
        repartidor = self.request.GET.get('repartidor', '').strip()
        if fecha:
            qs = qs.filter(fecha=fecha)
        if repartidor:
            qs = qs.filter(nombre_repartidor__icontains=repartidor)
        return qs

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except (OperationalError, ProgrammingError) as exc:
            text = str(exc).lower()
            _write_rendiciones_log('rendiciones_list_error.log', traceback.format_exc())
            if (
                'peoneta_id' in text
                or 'routes_rutadia' in text
                or 'rendiciones_rendicionreparto' in text
                or ("doesn't exist" in text and 'rendiciones_' in text)
            ):
                messages.error(
                    request,
                    'La base de datos del hosting esta desactualizada (faltan migraciones). '
                    'Ejecuta: python manage.py migrate',
                )
                return redirect('dashboard:index')
            raise
        except Exception:
            _write_rendiciones_log('rendiciones_list_error.log', traceback.format_exc())
            messages.error(
                request,
                'Ocurrio un error al abrir Rendiciones. Revisa tmp/rendiciones_list_error.log para el detalle.',
            )
            return redirect('dashboard:index')


class RendicionDetailView(LoginRequiredMixin, DetailView):
    model = RendicionReparto
    template_name = 'rendiciones/detail.html'
    context_object_name = 'rendicion'


class RendicionDeleteView(LoginRequiredMixin, DeleteView):
    model = RendicionReparto
    template_name = 'rendiciones/confirm_delete.html'
    success_url = reverse_lazy('rendiciones:list')


@login_required
def plantilla_excel(request):
    template_path = Path(__file__).resolve().parent.parent / 'plantilla_rendicion_reparto.xlsx'
    if not template_path.exists():
        raise Http404('No se encontró la plantilla Excel.')

    response = FileResponse(
        open(template_path, 'rb'),
        as_attachment=True,
        filename='plantilla_rendicion_reparto.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return response


@login_required
def rendicion_excel(request, pk):
    try:
        load_workbook = _load_openpyxl()
    except Exception:
        messages.error(
            request,
            'La exportacion Excel no esta disponible en este entorno. Falta instalar la dependencia openpyxl.',
        )
        return redirect('rendiciones:detail', pk=pk)

    rendicion = get_object_or_404(
        RendicionReparto.objects.select_related('ruta__conductor', 'ruta__empresa').prefetch_related(
            'creditos_documentos',
            'devoluciones_parciales',
            'creditos_confianza',
            'facturas_nulas_detalle',
            'depositos_transferencias',
        ),
        pk=pk,
    )

    template_path = Path(__file__).resolve().parent.parent / 'plantilla_rendicion_reparto.xlsx'
    if not template_path.exists():
        raise Http404('No se encontró la plantilla Excel.')

    wb = load_workbook(template_path)
    ws = wb.active

    def set_money(cell, value):
        ws[cell] = int(value or 0)
        ws[cell].number_format = '#,##0'

    ws['C4'] = rendicion.distribuidora or ''
    ws['C5'] = rendicion.nombre_repartidor or ''
    ws['C6'] = rendicion.nombre_peoneta or ''
    ws['J5'] = rendicion.fecha.strftime('%d-%m-%Y')

    set_money('E8', rendicion.total_consolidado)
    set_money('E9', rendicion.menos_items)
    set_money('E10', rendicion.total_dinero_recibir)

    items_a = list(rendicion.creditos_documentos.all())
    for row_idx, item in enumerate(items_a[:7], start=14):
        ws[f'A{row_idx}'] = item.numero_factura or ''
        ws[f'C{row_idx}'] = item.nombre_cliente or ''
        set_money(f'E{row_idx}', item.monto)
        ws[f'F{row_idx}'] = item.banco or ''
    set_money('E21', sum((it.monto or 0) for it in items_a))

    items_b = list(rendicion.devoluciones_parciales.all())
    for row_idx, item in enumerate(items_b[:7], start=14):
        ws[f'H{row_idx}'] = item.numero_factura or ''
        ws[f'I{row_idx}'] = item.motivo or ''
        set_money(f'K{row_idx}', item.monto)
    set_money('K21', sum((it.monto or 0) for it in items_b))

    items_c = list(rendicion.creditos_confianza.all())
    for row_idx, item in enumerate(items_c[:13], start=25):
        ws[f'A{row_idx}'] = item.numero_factura or ''
        ws[f'C{row_idx}'] = item.autoriza_credito or ''
        set_money(f'F{row_idx}', item.monto)
    set_money('F38', sum((it.monto or 0) for it in items_c))

    items_d = list(rendicion.facturas_nulas_detalle.all())
    for row_idx, pair_start in enumerate(range(0, min(len(items_d), 10), 2), start=25):
        left = items_d[pair_start]
        ws[f'H{row_idx}'] = left.numero_factura or ''
        set_money(f'I{row_idx}', left.monto)
        if pair_start + 1 < len(items_d):
            right = items_d[pair_start + 1]
            ws[f'J{row_idx}'] = right.numero_factura or ''
            set_money(f'K{row_idx}', right.monto)
    set_money('K30', sum((it.monto or 0) for it in items_d))

    items_e = list(rendicion.depositos_transferencias.all())
    max_items_e = min(len(items_e), 28)
    for row_idx, pair_start in enumerate(range(0, max_items_e, 2), start=34):
        left = items_e[pair_start]
        ws[f'H{row_idx}'] = left.numero_factura or ''
        set_money(f'I{row_idx}', left.monto)
        if pair_start + 1 < len(items_e):
            right = items_e[pair_start + 1]
            ws[f'J{row_idx}'] = right.numero_factura or ''
            set_money(f'K{row_idx}', right.monto)
    set_money('K48', sum((it.monto or 0) for it in items_e))

    ws['E40'] = rendicion.facturas_totales or 0
    ws['E41'] = rendicion.facturas_entregadas or 0
    ws['E42'] = rendicion.facturas_nulas or 0
    ws['E44'] = rendicion.kilometraje_inicial or 0
    ws['E45'] = rendicion.kilometraje_final or 0
    ws['E46'] = rendicion.total_kilometros_recorridos or 0

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f'rendicion_{rendicion.pk}_{rendicion.fecha.strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _initial_from_ruta(ruta):
    return {
        'ruta': ruta,
        'fecha': ruta.fecha,
        'nombre_repartidor': ruta.conductor_id,
        'nombre_peoneta': ruta.peoneta_id,
        'distribuidora': ruta.empresa_id,
        'total_consolidado': ruta.total_consolidado,
    }


@login_required
def rendicion_create(request):
    ruta_id = request.GET.get('ruta') or request.POST.get('ruta')
    initial = {}
    ruta = None
    if ruta_id:
        from routes.models import RutaDia
        ruta = get_object_or_404(RutaDia.objects.select_related('conductor', 'empresa'), pk=ruta_id)
        if hasattr(ruta, 'rendicion'):
            messages.info(request, 'La ruta ya tiene una rendición asociada.')
            return redirect('rendiciones:update', pk=ruta.rendicion.pk)
        initial = _initial_from_ruta(ruta)

    autocompletado_info = None
    if request.method == 'POST':
        form = RendicionRepartoForm(request.POST)
        clientes_ruta_sugeridos = get_clientes_ruta_nombres(ruta) if ruta else []
        if not ruta and request.POST.get('ruta'):
            from routes.models import RutaDia
            try:
                ruta = RutaDia.objects.select_related('conductor', 'empresa').get(pk=request.POST.get('ruta'))
            except RutaDia.DoesNotExist:
                ruta = None
        clientes_ruta_sugeridos = get_clientes_ruta_nombres(ruta) if ruta else []
        formsets = build_formsets(data=request.POST, instance=None, ruta=ruta, clientes_sugeridos=clientes_ruta_sugeridos)
        if form.is_valid() and all(fs.is_valid() for fs in formsets.values()):
            rendicion = form.save(commit=False)
            if not rendicion.nombre_repartidor:
                rendicion.nombre_repartidor = rendicion.ruta.conductor.get_full_name() or rendicion.ruta.conductor.username
            if not rendicion.total_consolidado:
                rendicion.total_consolidado = rendicion.ruta.total_consolidado
            rendicion.save()
            form.save_m2m()
            for fs in formsets.values():
                fs.instance = rendicion
                fs.save()
            _autocompletar_rendicion_desde_entregas(rendicion)
            rendicion.recalcular_totales()
            rendicion.save(update_fields=['menos_items', 'total_dinero_recibir'])
            messages.success(request, 'Rendición creada correctamente.')
            return redirect('rendiciones:detail', pk=rendicion.pk)
    else:
        form = RendicionRepartoForm(initial=initial)
        clientes_ruta_sugeridos = get_clientes_ruta_nombres(ruta) if ruta else []
        initial_formsets, autocompletado_info = _build_autocompletado_desde_ruta(ruta)
        formsets = build_formsets(
            instance=None,
            ruta=ruta,
            clientes_sugeridos=clientes_ruta_sugeridos,
            initial_data=initial_formsets,
        )

    ctx = {
        'titulo': 'Nueva rendición de reparto',
        'form': form,
        'clientes_ruta_sugeridos': clientes_ruta_sugeridos,
        'autocompletado_info': autocompletado_info,
        **formsets,
    }
    return render(request, 'rendiciones/form.html', ctx)


@login_required
def rendicion_update(request, pk):
    rendicion = get_object_or_404(RendicionReparto, pk=pk)
    ruta = rendicion.ruta
    clientes_ruta_sugeridos = get_clientes_ruta_nombres(ruta)

    if request.method == 'POST':
        form = RendicionRepartoForm(request.POST, instance=rendicion)
        formsets = build_formsets(data=request.POST, instance=rendicion, ruta=ruta, clientes_sugeridos=clientes_ruta_sugeridos)
        if form.is_valid() and all(fs.is_valid() for fs in formsets.values()):
            rendicion = form.save()
            for fs in formsets.values():
                fs.save()
            # En update no se regenera desde ruta para preservar cambios manuales
            # y evitar duplicados cuando se edita numero_factura.
            rendicion.recalcular_totales()
            rendicion.save(update_fields=['menos_items', 'total_dinero_recibir'])
            messages.success(request, 'Rendición actualizada correctamente.')
            return redirect('rendiciones:detail', pk=rendicion.pk)
    else:
        form = RendicionRepartoForm(instance=rendicion)
        formsets = build_formsets(instance=rendicion, ruta=ruta, clientes_sugeridos=clientes_ruta_sugeridos)

    ctx = {
        'titulo': 'Editar rendición de reparto',
        'form': form,
        'rendicion': rendicion,
        'clientes_ruta_sugeridos': clientes_ruta_sugeridos,
        **formsets,
    }
    return render(request, 'rendiciones/form.html', ctx)


@login_required
def rendicion_pdf(request, pk):
    try:
        letter, canvas = _load_reportlab()
    except Exception:
        messages.error(
            request,
            'La exportacion PDF no esta disponible en este entorno. Falta instalar la dependencia reportlab.',
        )
        return redirect('rendiciones:detail', pk=pk)

    rendicion = get_object_or_404(
        RendicionReparto.objects.select_related('ruta__conductor', 'ruta__empresa'),
        pk=pk,
    )

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    page_w, page_h = letter

    def draw_box(x, y_top, w, h):
        c.rect(x, y_top - h, w, h)

    def draw_hline(x, y_top, w):
        c.line(x, y_top, x + w, y_top)

    def draw_vline(x, y_top, h):
        c.line(x, y_top, x, y_top - h)

    def draw_text(x, y, txt, bold=False, size=8):
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', size)
        c.drawString(x, y, str(txt or ''))

    def money(v):
        return _money(v or Decimal('0'))

    def sum_qs(qs):
        return qs.aggregate(total=Sum('monto'))['total'] or Decimal('0')

    margin = 12
    gap = 8
    left_w = 308
    right_w = page_w - (margin * 2) - left_w - gap
    x_left = margin
    x_right = margin + left_w + gap
    y = page_h - 16

    draw_text(18, page_h - 20, 'PADACI', bold=True, size=12)
    draw_text(18, page_h - 36, 'SPA', bold=True, size=12)
    draw_text(page_w / 2 - 120, page_h - 26, 'RENDICION DE VALORES POR REPARTO', bold=True, size=11)

    y -= 36

    h_header = 58
    draw_box(x_left, y, left_w, h_header)
    row_h = 19
    for i in range(1, 3):
        draw_hline(x_left, y - i * row_h, left_w)
    draw_vline(x_left + 120, y, h_header)
    draw_text(x_left + 4, y - 14, 'Distribuidora', size=8.5)
    draw_text(x_left + 124, y - 14, rendicion.distribuidora, size=8.5)
    draw_text(x_left + 4, y - 33, 'Nombre del Repartidor', size=8.5)
    draw_text(x_left + 124, y - 33, rendicion.nombre_repartidor, size=8.5)
    draw_text(x_left + 4, y - 52, 'Nombre del Peoneta', size=8.5)
    draw_text(x_left + 124, y - 52, rendicion.nombre_peoneta, size=8.5)

    fecha_h = 20
    draw_box(x_right, y - 19, right_w * 0.85, fecha_h)
    draw_vline(x_right + 60, y - 19, fecha_h)
    draw_text(x_right + 20, y - 33, 'Fecha', size=10)
    draw_text(x_right + 64, y - 33, rendicion.fecha.strftime('%d-%m-%Y'), size=9)

    y -= h_header + 12

    h_sum1 = 48
    draw_box(x_left, y, left_w, h_sum1)
    draw_hline(x_left, y - 24, left_w)
    draw_vline(x_left + 185, y, h_sum1)
    draw_text(x_left + 4, y - 16, 'Total consolidado', bold=True, size=9)
    draw_text(x_left + 188, y - 16, money(rendicion.total_consolidado), size=9)
    draw_text(x_left + 4, y - 40, 'Menos Item (A,B,C,D,E)', bold=True, size=9)
    draw_text(x_left + 188, y - 40, money(rendicion.menos_items), size=9)

    h_sum2 = 96
    draw_box(x_right, y + 18, right_w, h_sum2)
    s_row_h = 19
    for i in range(1, 5):
        draw_hline(x_right, y + 18 - i * s_row_h, right_w)
    draw_vline(x_right + right_w - 76, y + 18, h_sum2)
    draw_text(x_right + 6, y + 3, 'Estacionamientos', size=8.5)
    draw_text(x_right + right_w - 72, y + 3, money(rendicion.estacionamientos), size=8.5)
    draw_text(x_right + 6, y - 16, 'Diferencia menos', size=8.5)
    draw_text(x_right + right_w - 72, y - 16, money(rendicion.diferencia_menos), size=8.5)
    draw_text(x_right + 6, y - 35, 'Diferencia más', size=8.5)
    draw_text(x_right + right_w - 72, y - 35, money(rendicion.diferencia_mas), size=8.5)
    draw_text(x_right + 6, y - 54, 'Total consolidado', bold=True, size=8.5)
    draw_text(x_right + right_w - 72, y - 54, money(rendicion.total_consolidado), size=8.5)
    draw_text(x_right + 6, y - 73, 'Total dinero a recibir', bold=True, size=8.5)
    draw_text(x_right + right_w - 72, y - 73, money(rendicion.total_dinero_recibir), bold=True, size=8.5)

    y -= 64

    h_total_recibir = 24
    draw_box(x_left, y, left_w, h_total_recibir)
    draw_vline(x_left + 185, y, h_total_recibir)
    draw_text(x_left + 4, y - 16, 'Total dinero a recibir', bold=True, size=9)
    draw_text(x_left + 188, y - 16, money(rendicion.total_dinero_recibir), bold=True, size=9)

    def draw_grid_section(x, y_top, w, title, headers, data_rows, col_widths, total_label=None, total_value=None, rows_visible=7):
        title_h = 20
        header_h = 20
        row_h_inner = 19
        total_h = 20 if total_label else 0
        h = title_h + header_h + rows_visible * row_h_inner + total_h
        draw_box(x, y_top, w, h)

        draw_hline(x, y_top - title_h, w)
        draw_text(x + w / 2 - (len(title) * 2.0), y_top - 14, title, size=8.5)

        draw_hline(x, y_top - title_h - header_h, w)
        acc = x
        for cw in col_widths[:-1]:
            acc += cw
            draw_vline(acc, y_top - title_h, h - title_h)

        header_y = y_top - title_h - 14
        tx = x + 4
        for i, hdr in enumerate(headers):
            draw_text(tx, header_y, hdr, size=8.5)
            tx += col_widths[i]

        base_y = y_top - title_h - header_h
        for i in range(1, rows_visible + 1):
            draw_hline(x, base_y - i * row_h_inner, w)

        rows = data_rows[:rows_visible]
        for idx, row in enumerate(rows):
            yy = base_y - (idx * row_h_inner) - 14
            tx = x + 4
            for col_idx, cell in enumerate(row):
                draw_text(tx, yy, cell, size=8)
                tx += col_widths[col_idx]

        if total_label:
            draw_hline(x, base_y - rows_visible * row_h_inner - total_h, w)
            draw_text(x + 6, base_y - rows_visible * row_h_inner - 14, total_label, bold=True, size=8.5)
            draw_text(x + w - 70, base_y - rows_visible * row_h_inner - 14, total_value, bold=True, size=8.5)

        return h

    rows_a = [[r.numero_factura, r.nombre_cliente, money(r.monto), r.banco] for r in rendicion.creditos_documentos.all() if any([r.numero_factura, r.nombre_cliente, r.monto, r.banco])]
    rows_b = [[r.numero_factura, r.motivo, money(r.monto)] for r in rendicion.devoluciones_parciales.all() if any([r.numero_factura, r.motivo, r.monto])]
    rows_c = [[r.numero_factura, r.autoriza_credito, money(r.monto)] for r in rendicion.creditos_confianza.all() if any([r.numero_factura, r.autoriza_credito, r.monto])]

    d_raw = [[r.numero_factura, money(r.monto)] for r in rendicion.facturas_nulas_detalle.all() if any([r.numero_factura, r.monto])]
    e_raw = [[r.numero_factura, money(r.monto)] for r in rendicion.depositos_transferencias.all() if any([r.numero_factura, r.monto])]

    rows_d = []
    for i in range(0, len(d_raw), 2):
        l = d_raw[i]
        r = d_raw[i + 1] if i + 1 < len(d_raw) else ['', '']
        rows_d.append([l[0], l[1], r[0], r[1]])

    rows_e = []
    for i in range(0, len(e_raw), 2):
        l = e_raw[i]
        r = e_raw[i + 1] if i + 1 < len(e_raw) else ['', '']
        rows_e.append([l[0], l[1], r[0], r[1]])

    y -= 18
    h_a = draw_grid_section(
        x_left, y, left_w,
        'Credito con cheques adjuntos / pago con nota de credito (A)',
        ['N° factura', 'Nombre Cliente', 'Monto', 'Banco'],
        rows_a,
        [75, 120, 75, 38],
        total_label='Total créditos con documentos',
        total_value=money(sum_qs(rendicion.creditos_documentos.all())),
        rows_visible=8,
    )
    h_b = draw_grid_section(
        x_right, y, right_w,
        'Devoluciones parciales (B)',
        ['N° Factura', 'Motivo', 'Monto'],
        rows_b,
        [80, 120, 73],
        total_label='Total devoluciones parciales',
        total_value=money(sum_qs(rendicion.devoluciones_parciales.all())),
        rows_visible=6,
    )

    y -= max(h_a, h_b) + 10
    h_c = draw_grid_section(
        x_left, y, left_w,
        'Creditos de confianza (solo firma) (C)',
        ['N° factura', 'Quien autoriza el crédito', 'Monto'],
        rows_c,
        [75, 169, 64],
        total_label='Total créditos de confianza',
        total_value=money(sum_qs(rendicion.creditos_confianza.all())),
        rows_visible=6,
    )
    h_d = draw_grid_section(
        x_right, y, right_w,
        'Facturas nulas (D)',
        ['N° factura', 'Monto', 'N° factura', 'Monto'],
        rows_d,
        [68, 52, 68, 52],
        total_label='Total nulas',
        total_value=money(sum_qs(rendicion.facturas_nulas_detalle.all())),
        rows_visible=3,
    )

    y -= max(h_c, h_d) + 8
    h_footer_left = 78
    draw_box(x_left, y, left_w, h_footer_left)
    for i in range(1, 6):
        draw_hline(x_left, y - i * 13, left_w)
    draw_vline(x_left + 185, y, h_footer_left)
    draw_text(x_left + 4, y - 10, 'Facturas Totales', size=8.5)
    draw_text(x_left + 188, y - 10, str(rendicion.facturas_totales), size=8.5)
    draw_text(x_left + 4, y - 23, 'Facturas Entregadas', size=8.5)
    draw_text(x_left + 188, y - 23, str(rendicion.facturas_entregadas), size=8.5)
    draw_text(x_left + 4, y - 36, 'Facturas Nulas', size=8.5)
    draw_text(x_left + 188, y - 36, str(rendicion.facturas_nulas), size=8.5)
    draw_text(x_left + 4, y - 49, 'Kilometraje Inicial', size=8.5)
    draw_text(x_left + 188, y - 49, str(rendicion.kilometraje_inicial), size=8.5)
    draw_text(x_left + 4, y - 62, 'Kilometraje Final', size=8.5)
    draw_text(x_left + 188, y - 62, str(rendicion.kilometraje_final), size=8.5)
    draw_text(x_left + 4, y - 75, 'Total Kilometros Recorridos', size=8.5)
    draw_text(x_left + 188, y - 75, str(rendicion.total_kilometros_recorridos), size=8.5)

    draw_grid_section(
        x_right, y + 48, right_w,
        'Depositos o Transferencias (E)',
        ['N° factura', 'Monto', 'N° factura', 'Monto'],
        rows_e,
        [68, 52, 68, 52],
        total_label='Total',
        total_value=money(sum_qs(rendicion.depositos_transferencias.all())),
        rows_visible=5,
    )

    c.showPage()
    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="rendicion_{rendicion.pk}.pdf"'
    return response
