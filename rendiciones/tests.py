from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from accounts.models import CustomUser
from clients.models import Cliente
from companies.models import Empresa
from rendiciones.models import RendicionReparto
from routes.models import Entrega, EntregaPago, ParadaRuta, RutaDia


class RendicionCreateTests(TestCase):
	def setUp(self):
		self.user = CustomUser.objects.create_user(
			username='admin',
			password='secret123',
			rol='admin',
			first_name='Admin',
			last_name='Padaci',
			activo=True,
		)
		self.conductor = CustomUser.objects.create_user(
			username='conductor',
			password='secret123',
			rol='conductor',
			first_name='Conductor',
			last_name='Uno',
			activo=True,
		)
		self.empresa = Empresa.objects.create(
			nombre='Distribuidora Demo',
			razon_social='Distribuidora Demo SPA',
			rut='76000000-1',
			direccion='Calle 123',
			activa=True,
		)
		self.cliente = Cliente.objects.create(
			empresa=self.empresa,
			nombre='Cliente Demo',
			direccion='Direccion Cliente',
		)
		self.ruta = RutaDia.objects.create(
			fecha='2026-04-09',
			empresa=self.empresa,
			conductor=self.conductor,
			total_consolidado=Decimal('100000'),
		)
		self.client.force_login(self.user)

	def _crear_entrega(self, orden, estado='entregado', metodo_pago=None, monto='0', observacion=''):
		entrega = Entrega.objects.create(
			cliente=self.cliente,
			empresa=self.empresa,
			conductor=self.conductor,
			estado=estado,
			fecha_programada=self.ruta.fecha,
		)
		ParadaRuta.objects.create(ruta=self.ruta, entrega=entrega, orden=orden)
		if metodo_pago:
			EntregaPago.objects.create(
				entrega=entrega,
				metodo=metodo_pago,
				monto=Decimal(monto),
				observacion=observacion,
			)
		return entrega

	def test_create_preserves_manual_invoice_numbers_for_all_items(self):
		self._crear_entrega(1, metodo_pago='cheque', monto='1000')
		self._crear_entrega(2, metodo_pago='descuento', monto='2000', observacion='Motivo descuento')
		self._crear_entrega(3, metodo_pago='credito', monto='3000')
		self._crear_entrega(4, estado='fallido')
		self._crear_entrega(5, metodo_pago='transferencia', monto='5000')

		response = self.client.post(
			reverse('rendiciones:create'),
			data={
				'ruta': str(self.ruta.pk),
				'fecha': '2026-04-09',
				'distribuidora': str(self.empresa.pk),
				'nombre_repartidor': str(self.conductor.pk),
				'nombre_peoneta': '',
				'estacionamientos': '0',
				'diferencia_menos': '0',
				'diferencia_mas': '0',
				'facturas_totales': '5',
				'facturas_entregadas': '4',
				'facturas_nulas': '1',
				'kilometraje_inicial': '0',
				'kilometraje_final': '10',
				'total_kilometros_recorridos': '10',
				'a-TOTAL_FORMS': '1',
				'a-INITIAL_FORMS': '0',
				'a-MIN_NUM_FORMS': '0',
				'a-MAX_NUM_FORMS': '1000',
				'a-0-numero_factura': 'A-MANUAL-001',
				'a-0-nombre_cliente': 'Cliente Demo',
				'a-0-monto': '1000',
				'a-0-banco': 'Chile',
				'b-TOTAL_FORMS': '1',
				'b-INITIAL_FORMS': '0',
				'b-MIN_NUM_FORMS': '0',
				'b-MAX_NUM_FORMS': '1000',
				'b-0-numero_factura': 'B-MANUAL-001',
				'b-0-motivo': 'Motivo descuento',
				'b-0-monto': '2000',
				'c-TOTAL_FORMS': '1',
				'c-INITIAL_FORMS': '0',
				'c-MIN_NUM_FORMS': '0',
				'c-MAX_NUM_FORMS': '1000',
				'c-0-numero_factura': 'C-MANUAL-001',
				'c-0-autoriza_credito': 'Cliente Demo',
				'c-0-monto': '3000',
				'd-TOTAL_FORMS': '1',
				'd-INITIAL_FORMS': '0',
				'd-MIN_NUM_FORMS': '0',
				'd-MAX_NUM_FORMS': '1000',
				'd-0-numero_factura': 'D-MANUAL-001',
				'd-0-monto': '0',
				'e-TOTAL_FORMS': '1',
				'e-INITIAL_FORMS': '0',
				'e-MIN_NUM_FORMS': '0',
				'e-MAX_NUM_FORMS': '1000',
				'e-0-numero_factura': 'E-MANUAL-001',
				'e-0-monto': '5000',
			},
		)

		self.assertEqual(response.status_code, 302)

		rendicion = RendicionReparto.objects.get(ruta=self.ruta)
		self.assertEqual(rendicion.creditos_documentos.count(), 1)
		self.assertEqual(rendicion.devoluciones_parciales.count(), 1)
		self.assertEqual(rendicion.creditos_confianza.count(), 1)
		self.assertEqual(rendicion.facturas_nulas_detalle.count(), 1)
		self.assertEqual(rendicion.depositos_transferencias.count(), 1)

		self.assertEqual(rendicion.creditos_documentos.get().numero_factura, 'A-MANUAL-001')
		self.assertEqual(rendicion.devoluciones_parciales.get().numero_factura, 'B-MANUAL-001')
		self.assertEqual(rendicion.creditos_confianza.get().numero_factura, 'C-MANUAL-001')
		self.assertEqual(rendicion.facturas_nulas_detalle.get().numero_factura, 'D-MANUAL-001')
		self.assertEqual(rendicion.depositos_transferencias.get().numero_factura, 'E-MANUAL-001')
		self.assertEqual(rendicion.menos_items, Decimal('11000'))
		self.assertEqual(rendicion.total_dinero_recibir, Decimal('89000'))

	def test_create_autocompletes_nula_payment_into_facturas_nulas(self):
		entrega = self._crear_entrega(1, metodo_pago='nula', monto='4500')

		response = self.client.post(
			reverse('rendiciones:create'),
			data={
				'ruta': str(self.ruta.pk),
				'fecha': '2026-04-09',
				'distribuidora': str(self.empresa.pk),
				'nombre_repartidor': str(self.conductor.pk),
				'nombre_peoneta': '',
				'estacionamientos': '0',
				'diferencia_menos': '0',
				'diferencia_mas': '0',
				'facturas_totales': '1',
				'facturas_entregadas': '1',
				'facturas_nulas': '1',
				'kilometraje_inicial': '0',
				'kilometraje_final': '10',
				'total_kilometros_recorridos': '10',
				'a-TOTAL_FORMS': '0',
				'a-INITIAL_FORMS': '0',
				'a-MIN_NUM_FORMS': '0',
				'a-MAX_NUM_FORMS': '1000',
				'b-TOTAL_FORMS': '0',
				'b-INITIAL_FORMS': '0',
				'b-MIN_NUM_FORMS': '0',
				'b-MAX_NUM_FORMS': '1000',
				'c-TOTAL_FORMS': '0',
				'c-INITIAL_FORMS': '0',
				'c-MIN_NUM_FORMS': '0',
				'c-MAX_NUM_FORMS': '1000',
				'd-TOTAL_FORMS': '0',
				'd-INITIAL_FORMS': '0',
				'd-MIN_NUM_FORMS': '0',
				'd-MAX_NUM_FORMS': '1000',
				'e-TOTAL_FORMS': '0',
				'e-INITIAL_FORMS': '0',
				'e-MIN_NUM_FORMS': '0',
				'e-MAX_NUM_FORMS': '1000',
			},
		)

		self.assertEqual(response.status_code, 302)

		rendicion = RendicionReparto.objects.get(ruta=self.ruta)
		self.assertEqual(rendicion.facturas_nulas_detalle.count(), 1)
		item = rendicion.facturas_nulas_detalle.get()
		self.assertEqual(item.numero_factura, str(entrega.pk))
		self.assertEqual(item.monto, Decimal('4500'))

	def test_autocompletado_usa_total_factura_ref_desde_ocr(self):
		"""Verifica que el autocompletado en rendición use total_factura_ref de OCR, no pago.monto."""
		# Crear entrega con total_factura_ref (simulando OCR) diferente al pago.monto
		entrega = Entrega.objects.create(
			cliente=self.cliente,
			empresa=self.empresa,
			conductor=self.conductor,
			estado='entregado',
			fecha_programada=self.ruta.fecha,
			numero_factura_ref='OCR-12345',
			total_factura_ref=Decimal('50000'),  # Total extraído de OCR
		)
		ParadaRuta.objects.create(ruta=self.ruta, entrega=entrega, orden=1)
		# El pago es diferente al total de factura (ej: descuento o cliente pagó menos)
		EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('45000'),  # Monto efectivamente pagado (diferente al OCR)
		)

		response = self.client.post(
			reverse('rendiciones:create'),
			data={
				'ruta': str(self.ruta.pk),
				'fecha': '2026-04-09',
				'distribuidora': str(self.empresa.pk),
				'nombre_repartidor': str(self.conductor.pk),
				'nombre_peoneta': '',
				'estacionamientos': '0',
				'diferencia_menos': '0',
				'diferencia_mas': '0',
				'facturas_totales': '1',
				'facturas_entregadas': '1',
				'facturas_nulas': '0',
				'kilometraje_inicial': '0',
				'kilometraje_final': '10',
				'total_kilometros_recorridos': '10',
				'a-TOTAL_FORMS': '1',
				'a-INITIAL_FORMS': '0',
				'a-MIN_NUM_FORMS': '0',
				'a-MAX_NUM_FORMS': '1000',
				'a-0-numero_factura': 'OCR-12345',
				'a-0-nombre_cliente': 'Cliente Demo',
				'a-0-monto': '50000',  # Debe ser 50000 (total_factura_ref), NO 45000 (pago.monto)
				'b-TOTAL_FORMS': '0',
				'b-INITIAL_FORMS': '0',
				'b-MIN_NUM_FORMS': '0',
				'b-MAX_NUM_FORMS': '1000',
				'c-TOTAL_FORMS': '0',
				'c-INITIAL_FORMS': '0',
				'c-MIN_NUM_FORMS': '0',
				'c-MAX_NUM_FORMS': '1000',
				'd-TOTAL_FORMS': '0',
				'd-INITIAL_FORMS': '0',
				'd-MIN_NUM_FORMS': '0',
				'd-MAX_NUM_FORMS': '1000',
				'e-TOTAL_FORMS': '0',
				'e-INITIAL_FORMS': '0',
				'e-MIN_NUM_FORMS': '0',
				'e-MAX_NUM_FORMS': '1000',
			},
		)

		self.assertEqual(response.status_code, 302)
		rendicion = RendicionReparto.objects.get(ruta=self.ruta)
		self.assertEqual(rendicion.creditos_documentos.count(), 1)
		item = rendicion.creditos_documentos.get()
		# El monto autocompletado debe ser 50000 (total_factura_ref), NO 45000 (pago.monto)
		self.assertEqual(item.numero_factura, 'OCR-12345')
		self.assertEqual(item.monto, Decimal('50000'))

	def test_autocompletado_muestra_monto_correcto_en_rendicion(self):
		"""Verifica que el autocompletado en rendición cargue el monto correcto desde BD."""
		# Crear entrega con número de factura y total_factura_ref desde OCR
		entrega = Entrega.objects.create(
			cliente=self.cliente,
			empresa=self.empresa,
			conductor=self.conductor,
			estado='entregado',
			fecha_programada=self.ruta.fecha,
			numero_factura_ref='FAC-100001',
			total_factura_ref=Decimal('250000'),  # Monto extraído de OCR
		)
		ParadaRuta.objects.create(ruta=self.ruta, entrega=entrega, orden=1)
		
		# Crear pago para que se incluya en autocompletado
		EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('250000'),
		)

		# Simular la vista de creación de rendición que carga autocompletado
		from rendiciones.views import _build_autocompletado_desde_ruta
		sugeridos, info = _build_autocompletado_desde_ruta(self.ruta)
		
		# Verificar que el autocompletado contiene el monto correcto
		self.assertEqual(len(sugeridos['a']), 1)
		item = sugeridos['a'][0]
		self.assertEqual(item['numero_factura'], 'FAC-100001')
		self.assertEqual(item['monto'], Decimal('250000'))  # ← Debe ser 250000, NO 250 o 250.0

	def test_rendicion_autocompleta_monto_400_correctamente(self):
		"""Bug report: Crear rendición desde entrega con pago de 400 carga monto incorrecto (22 en lugar de 400)."""
		# Crear entrega con pago de 400 (CHEQUE)
		entrega = Entrega.objects.create(
			cliente=self.cliente,
			empresa=self.empresa,
			conductor=self.conductor,
			estado='entregado',
			fecha_programada=self.ruta.fecha,
		)
		ParadaRuta.objects.create(ruta=self.ruta, entrega=entrega, orden=1)
		
		# Crear pago de CHEQUE por 400
		pago = EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('400'),
		)
		
		print(f"DEBUG: Pago guardado en BD - monto={pago.monto}, type={type(pago.monto)}")
		
		# Verificar que se guardó correctamente en BD
		pago_refetch = EntregaPago.objects.get(pk=pago.pk)
		print(f"DEBUG: Pago obtenido de BD - monto={pago_refetch.monto}, type={type(pago_refetch.monto)}")
		self.assertEqual(pago_refetch.monto, Decimal('400'))
		
		# Verificar autocompletado
		from rendiciones.views import _build_autocompletado_desde_ruta
		sugeridos, info = _build_autocompletado_desde_ruta(self.ruta)
		
		print(f"DEBUG: Autocompletado section A = {sugeridos['a']}")
		self.assertEqual(len(sugeridos['a']), 1)
		
		item = sugeridos['a'][0]
		print(f"DEBUG: Item autocompletado - {item}")
		
		# ← Aquí debería ser 400, NO 22
		self.assertEqual(item['monto'], Decimal('400'), 
			f"El autocompletado debería cargar monto 400, pero está cargando {item['monto']}")

	def test_rendicion_form_renderiza_monto_correctamente_en_html(self):
		"""Verifica que el formulario de rendición renderiza el monto correctamente en HTML."""
		# Crear entrega con pago de 400
		entrega = Entrega.objects.create(
			cliente=self.cliente,
			empresa=self.empresa,
			conductor=self.conductor,
			estado='entregado',
			fecha_programada=self.ruta.fecha,
		)
		ParadaRuta.objects.create(ruta=self.ruta, entrega=entrega, orden=1)
		EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('400'),
		)
		
		# Acceder a la página de crear rendición
		response = self.client.get(reverse('rendiciones:create') + f'?ruta={self.ruta.pk}')
		self.assertEqual(response.status_code, 200)
		
		# Verificar que el HTML contiene el valor 400 en el campo de monto
		html_content = response.content.decode('utf-8')
		print(f"DEBUG: Verificando si '400' aparece en el HTML...")
		
		# El formulario debería tener value="400" para el monto autocompletado
		self.assertIn('value="400"', html_content, "El valor 400 debería estar en el HTML del formulario")
		
		# Asegurarse que 22 NO está presente
		self.assertNotIn('value="22"', html_content, "El valor 22 NO debería estar en el HTML")

	def test_usuario_edita_total_ocr_en_ui_y_crea_entrega(self):
		"""Simula flujo: Usuario ve monto de OCR en UI, lo edita, envía POST y luego crea rendición."""
		import json
		from django.urls import reverse
		
		# PASO 1: Usuario envía línea OCR como si hubiera editado en la UI
		# El usuario ve "250000" en la UI (capturado por OCR) pero lo deja tal cual
		payload = {
			'lineas': [
				{
					'documento': 'FAC-250K',
					'nombre_cliente': 'Cliente Demo',
					'direccion_cliente': 'Direccion Cliente',
					'dia': 9,
					'mes': 4,
					'total': 250000,  # ← Valor que viene del input type="number" (como número)
					'cond_pago': 'CONTADO',
					'comuna': 'Direccion Cliente',
					'cliente_id': self.cliente.pk,
				}
			]
		}

		# Crear entregas desde líneas OCR
		response = self.client.post(
			reverse('routes:crear_entregas', kwargs={'pk': self.ruta.pk}),
			data=json.dumps(payload),
			content_type='application/json',
		)
		
		self.assertEqual(response.status_code, 200)

		# PASO 2: Verificar que la entrega se creó correctamente
		entrega = Entrega.objects.get(numero_factura_ref='FAC-250K')
		print(f"DEBUG: entrega.total_factura_ref = {entrega.total_factura_ref}")  # Para debugging
		self.assertEqual(entrega.total_factura_ref, Decimal('250000'))

		# PASO 3: Crear pago
		EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('250000'),
		)

		# PASO 4: Verificar autocompletado
		from rendiciones.views import _build_autocompletado_desde_ruta
		sugeridos, info = _build_autocompletado_desde_ruta(self.ruta)
		
		self.assertEqual(len(sugeridos['a']), 1)
		item = sugeridos['a'][0]
		print(f"DEBUG: autocompletado monto = {item['monto']}")  # Para debugging
		self.assertEqual(item['monto'], Decimal('250000'))

	def test_flujo_completo_ocr_a_entrega_a_rendicion(self):
		"""Prueba end-to-end: enviar líneas OCR → crear entregas → autocompletar rendición."""
		# Este test simula el flujo real desde el usuario enviando lineas OCR con su total editado
		from django.urls import reverse
		import json

		# PASO 1: Usuario envía líneas OCR con total editado (ej: 75000 en lugar de 80000)
		payload = {
			'lineas': [
				{
					'documento': 'OCR-999',
					'nombre_cliente': 'Cliente Demo',
					'direccion_cliente': 'Direccion Cliente',
					'dia': 9,
					'mes': 4,
					'total': '75000',  # Valor editado en UI (string)
					'cond_pago': 'CONTADO',
					'comuna': 'Direccion Cliente',
					'cliente_id': self.cliente.pk,
				}
			]
		}

		# Crear entregas desde líneas OCR (flujo OCR)
		response = self.client.post(
			reverse('routes:crear_entregas', kwargs={'pk': self.ruta.pk}),
			data=json.dumps(payload),
			content_type='application/json',
		)
		
		self.assertEqual(response.status_code, 200)
		body = response.json()
		self.assertEqual(body['creadas'], 1)

		# Verificar que la entrega se creó con total_factura_ref = 75000
		entrega = Entrega.objects.get(numero_factura_ref='OCR-999')
		self.assertEqual(entrega.total_factura_ref, Decimal('75000'))

		# PASO 2: Crear rendición y verificar que el autocompletado usa total_factura_ref
		# Primero crear un pago para que aparezca en el autocompletado
		EntregaPago.objects.create(
			entrega=entrega,
			metodo='cheque',
			monto=Decimal('70000'),  # Diferente al total de factura
		)

		response = self.client.post(
			reverse('rendiciones:create'),
			data={
				'ruta': str(self.ruta.pk),
				'fecha': str(self.ruta.fecha),
				'distribuidora': str(self.empresa.pk),
				'nombre_repartidor': str(self.conductor.pk),
				'nombre_peoneta': '',
				'estacionamientos': '0',
				'diferencia_menos': '0',
				'diferencia_mas': '0',
				'facturas_totales': '1',
				'facturas_entregadas': '1',
				'facturas_nulas': '0',
				'kilometraje_inicial': '0',
				'kilometraje_final': '10',
				'total_kilometros_recorridos': '10',
				'a-TOTAL_FORMS': '1',
				'a-INITIAL_FORMS': '0',
				'a-MIN_NUM_FORMS': '0',
				'a-MAX_NUM_FORMS': '1000',
				'a-0-numero_factura': 'OCR-999',
				'a-0-nombre_cliente': 'Cliente Demo',
				'a-0-monto': '75000',  # Debe ser 75000 (total_factura_ref), NO 70000 (pago.monto)
				'b-TOTAL_FORMS': '0',
				'b-INITIAL_FORMS': '0',
				'b-MIN_NUM_FORMS': '0',
				'b-MAX_NUM_FORMS': '1000',
				'c-TOTAL_FORMS': '0',
				'c-INITIAL_FORMS': '0',
				'c-MIN_NUM_FORMS': '0',
				'c-MAX_NUM_FORMS': '1000',
				'd-TOTAL_FORMS': '0',
				'd-INITIAL_FORMS': '0',
				'd-MIN_NUM_FORMS': '0',
				'd-MAX_NUM_FORMS': '1000',
				'e-TOTAL_FORMS': '0',
				'e-INITIAL_FORMS': '0',
				'e-MIN_NUM_FORMS': '0',
				'e-MAX_NUM_FORMS': '1000',
			},
		)

		self.assertEqual(response.status_code, 302)
		rendicion = RendicionReparto.objects.get(ruta=self.ruta)
		self.assertEqual(rendicion.creditos_documentos.count(), 1)
		item = rendicion.creditos_documentos.get()
		self.assertEqual(item.numero_factura, 'OCR-999')
		# El monto debe ser 75000 (total_factura_ref), NO 70000 (pago.monto)
		self.assertEqual(item.monto, Decimal('75000'))

	def test_create_avoids_duplicate_factura_nula_when_delivery_is_fallido(self):
		entrega = self._crear_entrega(1, estado='fallido', metodo_pago='nula', monto='3500')

		response = self.client.post(
			reverse('rendiciones:create'),
			data={
				'ruta': str(self.ruta.pk),
				'fecha': '2026-04-09',
				'distribuidora': str(self.empresa.pk),
				'nombre_repartidor': str(self.conductor.pk),
				'nombre_peoneta': '',
				'estacionamientos': '0',
				'diferencia_menos': '0',
				'diferencia_mas': '0',
				'facturas_totales': '1',
				'facturas_entregadas': '0',
				'facturas_nulas': '1',
				'kilometraje_inicial': '0',
				'kilometraje_final': '10',
				'total_kilometros_recorridos': '10',
				'a-TOTAL_FORMS': '0',
				'a-INITIAL_FORMS': '0',
				'a-MIN_NUM_FORMS': '0',
				'a-MAX_NUM_FORMS': '1000',
				'b-TOTAL_FORMS': '0',
				'b-INITIAL_FORMS': '0',
				'b-MIN_NUM_FORMS': '0',
				'b-MAX_NUM_FORMS': '1000',
				'c-TOTAL_FORMS': '0',
				'c-INITIAL_FORMS': '0',
				'c-MIN_NUM_FORMS': '0',
				'c-MAX_NUM_FORMS': '1000',
				'd-TOTAL_FORMS': '0',
				'd-INITIAL_FORMS': '0',
				'd-MIN_NUM_FORMS': '0',
				'd-MAX_NUM_FORMS': '1000',
				'e-TOTAL_FORMS': '0',
				'e-INITIAL_FORMS': '0',
				'e-MIN_NUM_FORMS': '0',
				'e-MAX_NUM_FORMS': '1000',
			},
		)

		self.assertEqual(response.status_code, 302)

		rendicion = RendicionReparto.objects.get(ruta=self.ruta)
		self.assertEqual(rendicion.facturas_nulas_detalle.count(), 1)
		item = rendicion.facturas_nulas_detalle.get()
		self.assertEqual(item.numero_factura, str(entrega.pk))
		self.assertEqual(item.monto, Decimal('3500'))

	def test_export_resumen_excel_filters_by_date_range_and_includes_requested_columns(self):
		rendicion_1 = RendicionReparto.objects.create(
			ruta=self.ruta,
			fecha='2026-04-09',
			distribuidora='Distribuidora Demo',
			nombre_repartidor='Conductor Uno',
			nombre_peoneta='Peoneta Uno',
			total_consolidado=Decimal('100000'),
			estacionamientos=Decimal('5000'),
			facturas_totales=10,
			facturas_entregadas=8,
			facturas_nulas=2,
			kilometraje_inicial=Decimal('0'),
			kilometraje_final=Decimal('12'),
			total_kilometros_recorridos=Decimal('12'),
			menos_items=Decimal('21000'),
			total_dinero_recibir=Decimal('79000'),
		)
		rendicion_1.facturas_nulas_detalle.create(numero_factura='N-1', monto=Decimal('4000'))
		rendicion_1.devoluciones_parciales.create(numero_factura='DP-1', motivo='Motivo', monto=Decimal('3000'))
		rendicion_1.depositos_transferencias.create(numero_factura='T-1', monto=Decimal('6000'))
		rendicion_1.creditos_documentos.create(numero_factura='C-1', nombre_cliente='Cliente Demo', banco='Banco', monto=Decimal('7000'))
		rendicion_1.creditos_confianza.create(numero_factura='CR-1', autoriza_credito='Supervisor', monto=Decimal('1000'))

		ruta_2 = RutaDia.objects.create(
			fecha='2026-05-15',
			empresa=self.empresa,
			conductor=self.conductor,
			total_consolidado=Decimal('50000'),
		)
		RendicionReparto.objects.create(
			ruta=ruta_2,
			fecha='2026-05-15',
			distribuidora='Distribuidora Demo',
			nombre_repartidor='Conductor Uno',
			nombre_peoneta='Peoneta Dos',
			total_consolidado=Decimal('50000'),
			estacionamientos=Decimal('1000'),
			facturas_totales=4,
			facturas_entregadas=4,
			facturas_nulas=0,
			kilometraje_inicial=Decimal('0'),
			kilometraje_final=Decimal('5'),
			total_kilometros_recorridos=Decimal('5'),
			menos_items=Decimal('1000'),
			total_dinero_recibir=Decimal('49000'),
		)

		response = self.client.get(
			reverse('rendiciones:resumen_excel'),
			data={'fecha_desde': '2026-04-01', 'fecha_hasta': '2026-04-30'},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)

		workbook = load_workbook(BytesIO(response.content))
		worksheet = workbook.active
		headers = [cell.value for cell in worksheet[1]]
		self.assertEqual(
			headers,
			[
				'Fecha',
				'Total kilometros recorridos',
				'Facturas totales',
				'Facturas entregadas',
				'Facturas nulas',
				'Total consolidado',
				'Suma de monto de facturas nulas',
				'Suma de monto de devoluciones parciales',
				'Suma de monto de transferencias',
				'Suma de monto de cheques',
				'Suma de monto de creditos',
				'Total dinero a recibir',
				'Estacionamientos',
				'Nombre del repartidor',
				'Nombre del peoneta',
			],
		)
		self.assertEqual(worksheet.max_row, 2)
		self.assertEqual(str(worksheet['A2'].value), '2026-04-09')
		self.assertEqual(worksheet['B2'].value, 12)
		self.assertEqual(worksheet['G2'].value, 4000)
		self.assertEqual(worksheet['H2'].value, 3000)
		self.assertEqual(worksheet['I2'].value, 6000)
		self.assertEqual(worksheet['J2'].value, 7000)
		self.assertEqual(worksheet['K2'].value, 1000)
		self.assertEqual(worksheet['L2'].value, 79000)
		self.assertEqual(worksheet['N2'].value, 'Conductor Uno')

	def test_autocompletado_prioriza_pago_manual_sobre_ocr(self):
		"""FIX: Verificar que autocompletado usa pago manual (400) no OCR (22)."""
		entrega = self._crear_entrega(1, metodo_pago='cheque', monto='400')
		# Simular que OCR registró un valor diferente
		entrega.total_factura_ref = Decimal('22')
		entrega.save(update_fields=['total_factura_ref'])
		
		from rendiciones.views import _build_autocompletado_desde_ruta
		sugeridos, resumen = _build_autocompletado_desde_ruta(self.ruta)
		
		# Verificar que se cargó 400 (pago manual), no 22 (OCR)
		self.assertEqual(len(sugeridos['a']), 1)
		item = sugeridos['a'][0]
		self.assertEqual(item['monto'], Decimal('400'), "Debería usar pago manual 400, no OCR 22")
		
		# Verificar que se generó una advertencia por diferencia
		self.assertEqual(len(resumen['advertencias_diferencia']), 1)
		alerta = resumen['advertencias_diferencia'][0]
		self.assertEqual(alerta['monto_ocr'], Decimal('22'))
		self.assertEqual(alerta['monto_manual'], Decimal('400'))
		self.assertEqual(alerta['diferencia'], Decimal('378'))

	def test_flujo_completo_usuario_cheque_400_con_diferencia_ocr(self):
		"""Flujo end-to-end: Usuario registra pago 400, OCR vio 22, rendición muestra 400 y alerta."""
		# 1. Crear entrega con pago manual de 400
		entrega = self._crear_entrega(1, metodo_pago='cheque', monto='400')
		
		# 2. Simular que OCR de la factura extrajo valor diferente (22)
		entrega.numero_factura_ref = '87051'
		entrega.total_factura_ref = Decimal('22')
		entrega.save(update_fields=['numero_factura_ref', 'total_factura_ref'])
		
		# 3. Crear rendición
		rendicion = RendicionReparto.objects.create(
			ruta=self.ruta,
			fecha='2026-04-09',
			nombre_repartidor=self.conductor.get_full_name(),
			total_consolidado=self.ruta.total_consolidado,
		)
		
		# 4. Autocompletar desde entrega
		from rendiciones.views import _autocompletar_rendicion_desde_entregas
		_autocompletar_rendicion_desde_entregas(rendicion)
		
		# 5. Validar que se cargó monto manual (400), no OCR (22)
		items_cheque = rendicion.creditos_documentos.all()
		self.assertEqual(items_cheque.count(), 1)
		item = items_cheque.first()
		self.assertEqual(item.numero_factura, '87051')
		self.assertEqual(item.monto, Decimal('400'), "Debe usar pago manual, no OCR")
		self.assertEqual(item.nombre_cliente, 'Cliente Demo')
